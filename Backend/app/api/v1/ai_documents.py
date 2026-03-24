# -*- coding: utf-8 -*-

"""
AI Documents API Router
FastAPI endpoints for AI document generation
TenderZen v3.6 - AI Features

WIJZIGINGEN v3.6 (2026-03-18):
- PATCH  /tenders/{tender_id}/milestones/{milestone_id}  — datum, tijd, status, verantwoordelijke
- POST   /tenders/{tender_id}/milestones                 — nieuwe milestone aanmaken
- DELETE /tenders/{tender_id}/milestones/{milestone_id}  — verwijderen
- ExtractPlanningRequest: model parameter (selecteerbaar vanuit UI: haiku/sonnet/opus)
- extract_planning: gebruikt gekozen model (haiku default, sonnet/opus optioneel)

WIJZIGINGEN v3.5 (2026-03-13):
- PATCH/DELETE/POST /tenders/{tender_id}/planning-taken
- POST /tenders/{tender_id}/populate-from-template
- generate-backplanning: categorie per taak via Claude

WIJZIGINGEN v3.3:
- GET milestones, planning-taken, checklist-items endpoints

WIJZIGINGEN v3.2:
- PDF direct naar Claude via Anthropic document API
- GET /tenders/{document_id}/akkoord + POST /tenders/{document_id}/downstream
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Query, Request
from supabase import Client
from app.services.ai_documents.ai_document_service import AIDocumentService
from app.services.tender_service import TenderService
from app.core.database import get_supabase_async
from pydantic import BaseModel
from datetime import datetime
from typing import Optional, List
import anthropic
import os
import uuid
import base64
import json
import re
import fitz
from app.config import settings
from fastapi.responses import StreamingResponse
import io
from app.utils.markdown_to_docx import convert_markdown_to_docx


MAX_PDF_DIRECT_SIZE = 20 * 1024 * 1024

router = APIRouter(prefix="/ai-documents", tags=["AI Documents"])

STORAGE_BUCKET = "ai-documents"

DOWNSTREAM_TEMPLATES = {
    "offerte": {
        "vult": ["tenderplanning", "checklist"],
        "overschrijft": False,
    },
    "rode_draad": {
        "vult": ["tenderplanning", "projectplanning", "checklist", "team"],
        "overschrijft": True,
    }
}


# ============================================
# HELPER FUNCTIONS
# ============================================

def get_user_id_from_request(request: Request) -> Optional[str]:
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return None
        token = auth_header.replace('Bearer ', '')
        parts = token.split('.')
        if len(parts) != 3:
            return None
        payload_b64 = parts[1]
        padding = 4 - len(payload_b64) % 4
        if padding != 4:
            payload_b64 += '=' * padding
        payload_json = base64.urlsafe_b64decode(payload_b64)
        payload = json.loads(payload_json)
        return payload.get('sub')
    except Exception as e:
        print(f"⚠️ Could not extract user_id from JWT: {e}")
        return None


def map_template_row(row):
    color_map = {
        'samenvatting': 'blue',
        'offerte': 'green',
        'rode_draad': 'purple',
        'versie1_inschrijving': 'orange',
        'win_check': 'success'
    }
    template_key = row.get("template_key")
    return {
        "id": row.get("id"),
        "template_key": template_key,
        "template_name": row.get("naam"),
        "template_icon": row.get("icon", "📄"),
        "beschrijving": row.get("beschrijving", ""),
        "prompt_template": row.get("prompt_template"),
        "default_prompt": row.get("default_prompt"),
        "priority": row.get("volgorde", 999),
        "color": color_map.get(template_key, "blue"),
        "estimated_time_minutes": row.get("estimated_duration_minutes", 10),
        "recommended_documents": row.get("required_files", []),
        "is_active": row.get("is_active", True),
        "is_beta": row.get("is_beta", False),
        "created_at": str(row.get("created_at", "")),
        "updated_at": str(row.get("updated_at", "")),
    }


def map_document_row(row):
    return {
        "id": row.get("id"),
        "tender_id": row.get("tender_id"),
        "template_key": row.get("template_key"),
        "status": row.get("status"),
        "progress": row.get("progress", 0),
        "document_content": row.get("document_content"),
        "generated_file_url": row.get("generated_file_url"),
        "claude_tokens_used": row.get("claude_tokens_used"),
        "is_goedgekeurd": row.get("is_goedgekeurd", False),
        "goedgekeurd_op": str(row.get("goedgekeurd_op", "")),
        "created_at": str(row.get("created_at", "")),
        "completed_at": str(row.get("completed_at", "")),
    }


def map_tender_document_row(row):
    return {
        "id": row.get("id"),
        "tender_id": row.get("tender_id"),
        "file_name": row.get("file_name"),
        "original_file_name": row.get("original_file_name"),
        "file_size": row.get("file_size"),
        "file_type": row.get("file_type"),
        "document_type": row.get("document_type"),
        "storage_path": row.get("storage_path"),
        "uploaded_by": row.get("uploaded_by"),
        "uploaded_at": str(row.get("uploaded_at", "")),
        "created_at": str(row.get("created_at", "")),
    }


def prepare_pdf_for_claude(file_bytes: bytes, filename: str) -> dict:
    pdf_base64 = base64.standard_b64encode(file_bytes).decode("utf-8")
    return {
        "type": "document",
        "source": {
            "type": "base64",
            "media_type": "application/pdf",
            "data": pdf_base64,
        },
        "title": filename,
        "citations": {"enabled": False},
    }


def extract_pdf_text_fallback(file_bytes: bytes, max_chars: int = 60000) -> str:
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        text_parts = []
        total_chars = 0
        for page_num in range(len(doc)):
            if total_chars >= max_chars:
                text_parts.append(f"\n[... document ingekort na {max_chars} tekens ...]")
                break
            page = doc[page_num]
            page_text = page.get_text()
            remaining = max_chars - total_chars
            if len(page_text) > remaining:
                text_parts.append(page_text[:remaining])
                total_chars = max_chars
            else:
                text_parts.append(page_text)
                total_chars += len(page_text)
        doc.close()
        return "\n".join(text_parts).strip()
    except Exception as e:
        print(f"⚠️ PDF tekst extractie mislukt: {e}")
        return ""


def extract_word_text(file_bytes: bytes, max_chars: int = 60000) -> str:
    try:
        import docx
        import io
        doc = docx.Document(io.BytesIO(file_bytes))
        parts = []
        total_chars = 0
        for element in doc.element.body:
            if total_chars >= max_chars:
                parts.append("[... document ingekort ...]")
                break
            tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag
            if tag == 'p':
                from docx.oxml.ns import qn
                tekst = ''.join(r.text for r in element.findall('.//' + qn('w:t')))
                if tekst.strip():
                    parts.append(tekst)
                    total_chars += len(tekst)
            elif tag == 'tbl':
                from docx.oxml.ns import qn
                for rij in element.findall('.//' + qn('w:tr')):
                    cellen = []
                    for cel in rij.findall('.//' + qn('w:tc')):
                        cel_tekst = ''.join(
                            t.text for t in cel.findall('.//' + qn('w:t')) if t.text
                        )
                        cellen.append(cel_tekst.strip())
                    rij_tekst = ' | '.join(cellen)
                    if rij_tekst.strip():
                        parts.append(rij_tekst)
                        total_chars += len(rij_tekst)
        return '\n'.join(parts).strip()
    except Exception as e:
        print(f"⚠️ Word tekst extractie mislukt: {e}")
        return ""


def extract_excel_text(file_bytes: bytes, max_chars: int = 40000) -> str:
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        parts = []
        total_chars = 0
        for sheet_name in wb.sheetnames:
            if total_chars >= max_chars:
                parts.append("[... werkmap ingekort ...]")
                break
            ws = wb[sheet_name]
            parts.append(f"### Werkblad: {sheet_name}")
            for rij in ws.iter_rows():
                if total_chars >= max_chars:
                    parts.append("[... werkblad ingekort ...]")
                    break
                cellen = [str(cel.value) if cel.value is not None else '' for cel in rij]
                if any(c.strip() for c in cellen):
                    rij_tekst = ' | '.join(cellen)
                    parts.append(rij_tekst)
                    total_chars += len(rij_tekst)
        wb.close()
        return '\n'.join(parts).strip()
    except Exception as e:
        print(f"⚠️ Excel tekst extractie mislukt: {e}")
        return ""


def fetch_document_from_storage(db: Client, storage_path: str) -> Optional[bytes]:
    try:
        response = db.storage.from_(STORAGE_BUCKET).download(storage_path)
        return response
    except Exception as e:
        print(f"⚠️ Storage download mislukt voor {storage_path}: {e}")
        return None


# ============================================
# MARKDOWN PARSER
# ============================================

def parse_rode_draad_markdown(markdown: str) -> dict:
    result = {
        "tenderplanning": [],
        "projectplanning": [],
        "checklist": [],
        "team": []
    }
    sections = re.split(r'\n(?=## )', markdown)
    for section in sections:
        lines = section.strip().split('\n')
        if not lines:
            continue
        header = lines[0].lower()
        if 'tenderplanning' in header or 'planning (extern)' in header:
            for line in lines[1:]:
                if '|' in line and '---' not in line and 'mijlpaal' not in line.lower():
                    cols = [c.strip() for c in line.split('|') if c.strip()]
                    if len(cols) >= 2:
                        result["tenderplanning"].append({
                            "mijlpaal": cols[0],
                            "datum_tekst": cols[1],
                            "datum": _parse_date(cols[1])
                        })
        elif 'projectplanning' in header or 'planning (intern)' in header:
            for line in lines[1:]:
                if '|' in line and '---' not in line and 'taak' not in line.lower():
                    cols = [c.strip() for c in line.split('|') if c.strip()]
                    if len(cols) >= 2:
                        result["projectplanning"].append({
                            "taak_naam": cols[0],
                            "verantwoordelijke": cols[1] if len(cols) > 1 else "",
                            "deadline_tekst": cols[2] if len(cols) > 2 else "",
                            "deadline": _parse_date(cols[2]) if len(cols) > 2 else None
                        })
        elif 'checklist' in header or 'inlever' in header:
            for line in lines[1:]:
                line = line.strip()
                if line.startswith('-') or line.startswith('*'):
                    item_text = line.lstrip('-* ').strip()
                    is_verplicht = '(optioneel)' not in item_text.lower()
                    item_text = re.sub(r'\(verplicht\)|\(optioneel\)', '', item_text, flags=re.IGNORECASE).strip()
                    if item_text:
                        result["checklist"].append({
                            "taak_naam": item_text,
                            "is_verplicht": is_verplicht,
                            "sectie": "Inleverdocumenten"
                        })
        elif 'team' in header:
            for line in lines[1:]:
                if '|' in line and '---' not in line and 'naam' not in line.lower():
                    cols = [c.strip() for c in line.split('|') if c.strip()]
                    if len(cols) >= 2:
                        result["team"].append({
                            "naam": cols[0],
                            "rol": cols[1] if len(cols) > 1 else ""
                        })
    return result


def _parse_date(date_str: str) -> Optional[str]:
    if not date_str or date_str in ['-', 'n.v.t.', 'Niet opgegeven', '']:
        return None
    formats = ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d %B %Y', '%d %b %Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ============================================
# AI TEMPLATES ENDPOINTS
# ============================================

@router.get("/health")
async def health_check():
    return {'success': True, 'service': 'AI Documents', 'status': 'healthy'}


@router.get("/templates")
async def get_templates(
    only_active: bool = Query(True),
    db: Client = Depends(get_supabase_async)
):
    service = AIDocumentService(db)
    templates = await service.get_all_templates(only_active=only_active)
    mapped = [map_template_row(t) for t in templates]
    return {'success': True, 'templates': mapped, 'total': len(mapped)}


@router.get("/templates/{template_key}")
async def get_template(
    template_key: str,
    db: Client = Depends(get_supabase_async)
):
    service = AIDocumentService(db)
    template = await service.get_template_by_key(template_key)
    if not template:
        raise HTTPException(status_code=404, detail="Template niet gevonden")
    return {'success': True, 'template': map_template_row(template)}


@router.get("/tenders/{tender_id}/ai-documents")
async def get_tender_ai_documents(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    service = AIDocumentService(db)
    docs = await service.get_documents_for_tender(tender_id)
    mapped = [map_document_row(d) for d in docs]
    return {'success': True, 'documents': mapped, 'total': len(mapped)}


@router.get("/ai-documents/{document_id}")
async def get_ai_document(
    document_id: str,
    db: Client = Depends(get_supabase_async)
):
    service = AIDocumentService(db)
    doc = await service.get_document_by_id(document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document niet gevonden")
    return {'success': True, 'document': map_document_row(doc)}


# ============================================
# TENDER DOCUMENTS UPLOAD ENDPOINTS
# ============================================

@router.get("/tenders/{tender_id}/generated")
async def get_generated_documents(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('ai_documents') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False) \
            .order('created_at', desc=True) \
            .execute()
        documents = result.data or []
        return {'success': True, 'documents': documents, 'total': len(documents)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tenders/{tender_id}/brondocumenten")
async def get_brondocumenten(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        docs_result = db.table('tender_documents') \
            .select('id, original_file_name, file_name, file_type, file_size, document_type, uploaded_at, storage_path') \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False) \
            .order('uploaded_at', desc=True) \
            .execute()

        brondocumenten = []
        for doc in (docs_result.data or []):
            brondocumenten.append({
                "id": doc.get("id"),
                "naam": doc.get("original_file_name") or doc.get("file_name"),
                "type": "upload",
                "file_type": doc.get("file_type"),
                "file_size": doc.get("file_size"),
                "document_type": doc.get("document_type"),
                "storage_path": doc.get("storage_path"),
                "datum": str(doc.get("uploaded_at", "")),
                "aangevinkt": True
            })

        tender_result = db.table('tenders') \
            .select('smart_import_id') \
            .eq('id', tender_id) \
            .single() \
            .execute()

        if tender_result.data and tender_result.data.get('smart_import_id'):
            smart_import_id = tender_result.data['smart_import_id']
            si_result = db.table('smart_imports') \
                .select('id, uploaded_files, created_at, status') \
                .eq('id', smart_import_id) \
                .single() \
                .execute()
            if si_result.data and si_result.data.get('status') == 'completed':
                uploaded_files = si_result.data.get('uploaded_files', [])
                for f in uploaded_files:
                    brondocumenten.insert(0, {
                        "id": f"smart_import_{si_result.data['id']}",
                        "naam": f.get("original_name") or f.get("name", "Smart Import document"),
                        "type": "smart_import",
                        "file_type": f.get("type", "application/pdf"),
                        "file_size": f.get("size"),
                        "document_type": "aanbestedingsleidraad",
                        "storage_path": f.get("storage_path"),
                        "datum": str(si_result.data.get('created_at', "")),
                        "aangevinkt": True
                    })

        return {'success': True, 'brondocumenten': brondocumenten, 'total': len(brondocumenten)}
    except Exception as e:
        print(f"❌ Error fetching brondocumenten: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenders/{tender_id}/documents/upload")
async def upload_tender_document(
    tender_id: str,
    request: Request,
    document_type: str = Form(...),
    file: UploadFile = File(...),
    db: Client = Depends(get_supabase_async)
):
    try:
        max_size = 10 * 1024 * 1024
        file_content = await file.read()
        if len(file_content) > max_size:
            raise HTTPException(status_code=400, detail=f"File too large. Max 10MB.")

        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")

        tenderbureau_id = tender_result.data['tenderbureau_id']
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        storage_path = f"tenders/{tender_id}/{unique_filename}"

        try:
            db.storage.from_(STORAGE_BUCKET).upload(
                storage_path,
                file_content,
                {'content-type': file.content_type or 'application/octet-stream', 'cache-control': '3600', 'upsert': 'false'}
            )
        except Exception as storage_error:
            raise HTTPException(status_code=500, detail=f"Storage upload failed: {str(storage_error)}")

        document_data = {
            'tender_id': tender_id,
            'tenderbureau_id': tenderbureau_id,
            'file_name': unique_filename,
            'original_file_name': file.filename,
            'file_size': len(file_content),
            'file_type': file.content_type or 'application/octet-stream',
            'storage_path': storage_path,
            'document_type': document_type,
            'uploaded_by': get_user_id_from_request(request)
        }

        db_result = db.table('tender_documents').insert(document_data).execute()
        if not db_result.data:
            raise HTTPException(status_code=500, detail="Failed to save document metadata")

        return {
            'success': True,
            'document': map_tender_document_row(db_result.data[0]),
            'message': 'Document successfully uploaded'
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error uploading document: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tenders/{tender_id}/documents")
async def get_tender_documents(
    tender_id: str,
    document_type: Optional[str] = None,
    db: Client = Depends(get_supabase_async)
):
    try:
        query = db.table('tender_documents') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False) \
            .order('uploaded_at', desc=True)
        if document_type:
            query = query.eq('document_type', document_type)
        result = query.execute()
        documents = [map_tender_document_row(row) for row in result.data]
        return {'success': True, 'documents': documents, 'total': len(documents)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tenders/{tender_id}/documents/{document_id}")
async def get_tender_document(
    tender_id: str,
    document_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('tender_documents') \
            .select('*') \
            .eq('id', document_id) \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False) \
            .single() \
            .execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")
        return {'success': True, 'document': map_tender_document_row(result.data)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/tenders/{tender_id}/documents/{document_id}")
async def delete_tender_document(
    tender_id: str,
    document_id: str,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        doc_result = db.table('tender_documents') \
            .select('*') \
            .eq('id', document_id) \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False) \
            .single() \
            .execute()
        if not doc_result.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")

        db.table('tender_documents') \
            .update({
                'is_deleted': True,
                'deleted_at': datetime.now().isoformat(),
                'deleted_by': get_user_id_from_request(request)
            }) \
            .eq('id', document_id) \
            .execute()
        return {'success': True, 'message': 'Document successfully deleted'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# PROMPT MANAGEMENT ENDPOINTS
# ============================================

@router.post("/tenders/{tender_id}/fill-prompt/{template_key}")
async def fill_prompt_template(
    tender_id: str,
    template_key: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        tender_result = db.table('tenders').select('*').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")

        tender = tender_result.data
        tenderbureau_id = tender.get('tenderbureau_id')

        prompt_result = db.table('ai_prompts') \
            .select('*') \
            .eq('template_key', template_key) \
            .eq('status', 'active') \
            .or_(f'tenderbureau_id.eq.{tenderbureau_id},tenderbureau_id.is.null') \
            .order('tenderbureau_id', desc=False) \
            .limit(1) \
            .execute()

        if not prompt_result.data:
            raise HTTPException(status_code=404, detail=f"Geen actieve prompt voor '{template_key}'")

        prompt = prompt_result.data[0]
        prompt_content = prompt.get('prompt_content', '')

        docs_result = db.table('tender_documents').select('*').eq('tender_id', tender_id).eq('is_deleted', False).execute()
        documents = docs_result.data or []

        variables = {
            'tender_naam': tender.get('naam', ''),
            'tender_nummer': tender.get('tender_nummer', ''),
            'opdrachtgever': tender.get('opdrachtgever', ''),
            'aanbestedende_dienst': tender.get('aanbestedende_dienst') or tender.get('opdrachtgever', ''),
            'locatie': tender.get('locatie', 'Niet opgegeven'),
            'tender_waarde': f"€ {tender.get('tender_waarde', 0):,.2f}" if tender.get('tender_waarde') else 'Niet opgegeven',
            'deadline': str(tender.get('deadline_indiening', 'Niet opgegeven')),
            'omschrijving': tender.get('omschrijving', ''),
            'fase': tender.get('fase', 'onbekend'),
            'status': tender.get('status', 'onbekend'),
        }

        if documents:
            doc_list = [f"- {d.get('original_file_name', 'Document')}" for d in documents]
            variables['documenten_lijst'] = '\n'.join(doc_list)
            variables['aantal_documenten'] = str(len(documents))
        else:
            variables['documenten_lijst'] = '(Nog geen documenten geüpload)'
            variables['aantal_documenten'] = '0'

        filled_prompt = prompt_content
        for key, value in variables.items():
            filled_prompt = filled_prompt.replace(f'{{{{{key}}}}}', str(value))

        return {
            'success': True,
            'filled_prompt': filled_prompt,
            'variables': variables,
            'prompt_info': {
                'id': prompt.get('id'),
                'title': prompt.get('prompt_title'),
                'version': prompt.get('version'),
                'is_bureau_specific': prompt.get('tenderbureau_id') is not None
            },
            'template_key': template_key,
            'tender_id': tender_id
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prompts")
async def get_prompts(
    template_key: Optional[str] = None,
    status: Optional[str] = None,
    db: Client = Depends(get_supabase_async)
):
    try:
        query = db.table('ai_prompts').select('*')
        if template_key:
            query = query.eq('template_key', template_key)
        if status:
            query = query.eq('status', status)
        query = query.order('template_key').order('version', desc=True)
        result = query.execute()
        return {'success': True, 'prompts': result.data, 'total': len(result.data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prompts/active/{template_key}")
async def get_active_prompt(
    template_key: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('ai_prompts') \
            .select('*') \
            .eq('template_key', template_key) \
            .eq('status', 'active') \
            .is_('tenderbureau_id', 'null') \
            .single() \
            .execute()
        if not result.data:
            raise HTTPException(status_code=404, detail=f"Geen actieve prompt voor '{template_key}'")
        return {'success': True, 'prompt': result.data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prompts/{prompt_id}")
async def get_prompt(prompt_id: str, db: Client = Depends(get_supabase_async)):
    try:
        result = db.table('ai_prompts').select('*').eq('id', prompt_id).single().execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Prompt niet gevonden")
        return {'success': True, 'prompt': result.data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/prompts")
async def create_prompt(request: Request, prompt_data: dict, db: Client = Depends(get_supabase_async)):
    try:
        user_id = get_user_id_from_request(request)
        template_key = prompt_data.get('template_key')
        tenderbureau_id = prompt_data.get('tenderbureau_id')

        version_query = db.table('ai_prompts').select('version').eq('template_key', template_key)
        if tenderbureau_id:
            version_query = version_query.eq('tenderbureau_id', tenderbureau_id)
        else:
            version_query = version_query.is_('tenderbureau_id', 'null')

        version_result = version_query.order('version', desc=True).limit(1).execute()
        next_version = (version_result.data[0]['version'] + 1) if version_result.data else 1

        new_prompt = {
            'template_key': template_key,
            'tenderbureau_id': tenderbureau_id,
            'version': next_version,
            'prompt_title': prompt_data.get('prompt_title'),
            'prompt_content': prompt_data.get('prompt_content'),
            'variables': prompt_data.get('variables', []),
            'status': 'draft',
            'description': prompt_data.get('description'),
            'created_by': user_id
        }
        result = db.table('ai_prompts').insert(new_prompt).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Failed to create prompt")
        return {'success': True, 'prompt': result.data[0], 'message': f'Prompt versie {next_version} aangemaakt'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/prompts/{prompt_id}")
async def update_prompt(prompt_id: str, request: Request, updates: dict, db: Client = Depends(get_supabase_async)):
    try:
        current = db.table('ai_prompts').select('*').eq('id', prompt_id).single().execute()
        if not current.data:
            raise HTTPException(status_code=404, detail="Prompt niet gevonden")
        if current.data.get('status') != 'draft':
            raise HTTPException(status_code=400, detail="Alleen draft prompts kunnen bewerkt worden.")
        update_data = {
            'prompt_title': updates.get('prompt_title', current.data.get('prompt_title')),
            'prompt_content': updates.get('prompt_content', current.data.get('prompt_content')),
            'description': updates.get('description', current.data.get('description')),
            'variables': updates.get('variables', current.data.get('variables')),
            'updated_by': get_user_id_from_request(request)
        }
        result = db.table('ai_prompts').update(update_data).eq('id', prompt_id).execute()
        return {'success': True, 'prompt': result.data[0] if result.data else {}, 'message': 'Prompt bijgewerkt'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/prompts/{prompt_id}/activate")
async def activate_prompt(prompt_id: str, request: Request, db: Client = Depends(get_supabase_async)):
    try:
        prompt_result = db.table('ai_prompts').select('*').eq('id', prompt_id).single().execute()
        if not prompt_result.data:
            raise HTTPException(status_code=404, detail="Prompt niet gevonden")
        prompt = prompt_result.data
        template_key = prompt.get('template_key')
        tenderbureau_id = prompt.get('tenderbureau_id')
        user_id = get_user_id_from_request(request)

        deactivate_query = db.table('ai_prompts') \
            .update({'status': 'archived', 'updated_by': user_id}) \
            .eq('template_key', template_key) \
            .eq('status', 'active')
        if tenderbureau_id:
            deactivate_query = deactivate_query.eq('tenderbureau_id', tenderbureau_id)
        else:
            deactivate_query = deactivate_query.is_('tenderbureau_id', 'null')
        deactivate_query.execute()

        activate_result = db.table('ai_prompts') \
            .update({
                'status': 'active',
                'activated_at': datetime.now().isoformat(),
                'activated_by': user_id,
                'updated_by': user_id
            }) \
            .eq('id', prompt_id) \
            .execute()
        return {
            'success': True,
            'prompt': activate_result.data[0] if activate_result.data else {},
            'message': f"Prompt versie {prompt.get('version')} geactiveerd"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# DIRECTE AI GENERATIE
# ============================================

class GenerateDocumentRequest(BaseModel):
    template_key: str
    model: str = "claude-sonnet-4-6"
    brondocument_ids: Optional[List[str]] = None


@router.post("/tenders/{tender_id}/generate-document")
async def generate_document_for_tender(
    tender_id: str,
    body: GenerateDocumentRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    print(f"🤖 Genereer: tender={tender_id}, template={body.template_key}, model={body.model}")
    try:
        tender_result = db.table('tenders').select('*').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")
        tender = tender_result.data
        tenderbureau_id = tender.get('tenderbureau_id')

        prompts_result = db.table('ai_prompts') \
            .select('*') \
            .eq('template_key', body.template_key) \
            .eq('status', 'active') \
            .execute()

        if not prompts_result.data:
            raise HTTPException(status_code=404, detail=f"Geen actieve prompt voor '{body.template_key}'")

        prompts = prompts_result.data
        prompt_record = (
            next((p for p in prompts if p.get('tenderbureau_id') is None), None)
            or prompts[0]
        )
        if not prompt_record:
            raise HTTPException(status_code=404, detail="Geen bruikbare prompt gevonden")

        docs_query = db.table('tender_documents') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .eq('is_deleted', False)
        if body.brondocument_ids:
            docs_query = docs_query.in_('id', body.brondocument_ids)
        docs_result = docs_query.execute()
        documents = docs_result.data or []

        pdf_content_blocks = []
        fallback_teksten = []
        doc_namen_lijst_parts = []

        for doc in documents:
            storage_path = doc.get('storage_path')
            file_type = doc.get('file_type', '').lower()
            file_size = doc.get('file_size', 0) or 0
            original_name = doc.get('original_file_name') or doc.get('file_name', 'Document')
            naam_lower = original_name.lower()
            doc_namen_lijst_parts.append(f"- {original_name}")

            if not storage_path:
                continue

            file_bytes = fetch_document_from_storage(db, storage_path)
            if not file_bytes:
                continue

            is_pdf = 'pdf' in file_type or naam_lower.endswith('.pdf')
            is_word = 'wordprocessingml' in file_type or 'msword' in file_type \
                      or naam_lower.endswith('.docx') or naam_lower.endswith('.doc')
            is_excel = 'spreadsheetml' in file_type or 'excel' in file_type \
                       or naam_lower.endswith('.xlsx') or naam_lower.endswith('.xls')
            is_groot = file_size > MAX_PDF_DIRECT_SIZE

            if is_pdf and not is_groot:
                pdf_content_blocks.append(prepare_pdf_for_claude(file_bytes, original_name))
            elif is_pdf and is_groot:
                tekst = extract_pdf_text_fallback(file_bytes, max_chars=60000)
                fallback_teksten.append(f"=== {original_name} (PDF — tekst-extractie) ===\n{tekst or '(geen tekst)'}\n===")
            elif is_word:
                tekst = extract_word_text(file_bytes, max_chars=60000)
                fallback_teksten.append(f"=== {original_name} (Word document) ===\n{tekst or '(geen tekst)'}\n===")
            elif is_excel:
                tekst = extract_excel_text(file_bytes, max_chars=40000)
                fallback_teksten.append(f"=== {original_name} (Excel werkmap) ===\n{tekst or '(geen data)'}\n===")
            else:
                fallback_teksten.append(f"=== {original_name} ===\n(Bestandstype niet ondersteund)\n===")

        doc_namen_lijst = '\n'.join(doc_namen_lijst_parts) or '(Nog geen documenten geüpload)'
        fallback_tekst_blok = '\n\n'.join(fallback_teksten) if fallback_teksten else ''

        variables = {
            'tender_naam': tender.get('naam', ''),
            'tender_nummer': tender.get('tender_nummer', ''),
            'opdrachtgever': tender.get('opdrachtgever', ''),
            'aanbestedende_dienst': tender.get('aanbestedende_dienst') or tender.get('opdrachtgever', ''),
            'locatie': tender.get('locatie', 'Niet opgegeven'),
            'tender_waarde': str(tender.get('tender_waarde', 'Niet opgegeven')),
            'deadline': str(tender.get('deadline_indiening', 'Niet opgegeven')),
            'omschrijving': tender.get('omschrijving', ''),
            'documenten_lijst': doc_namen_lijst,
            'aantal_documenten': str(len(documents)),
            'documenten_inhoud': fallback_tekst_blok,
        }

        prompt_content = prompt_record.get('prompt_content', '')
        for key, value in variables.items():
            prompt_content = prompt_content.replace(f'{{{{{key}}}}}', str(value))

        model_map = {
            "haiku": "claude-haiku-4-5-20251001",
            "sonnet": "claude-sonnet-4-6",
            "opus": "claude-opus-4-6",
            "claude-haiku-4-5-20251001": "claude-haiku-4-5-20251001",
            "claude-sonnet-4-6": "claude-sonnet-4-6",
            "claude-opus-4-6": "claude-opus-4-6",
        }
        selected_model = model_map.get(body.model, "claude-sonnet-4-6")

        message_content = []
        if pdf_content_blocks:
            message_content.extend(pdf_content_blocks)
        message_content.append({"type": "text", "text": prompt_content})

        ai_client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        message = ai_client.messages.create(
            model=selected_model,
            max_tokens=8192,
            messages=[{"role": "user", "content": message_content}]
        )

        generated_text = message.content[0].text
        tokens_used = message.usage.input_tokens + message.usage.output_tokens

        doc_data = {
            'tender_id': tender_id,
            'tenderbureau_id': tenderbureau_id,
            'template_key': body.template_key,
            'status': 'completed',
            'progress': 100,
            'document_content': generated_text,
            'prompt_used': prompt_content,
            'input_data': {'brondocument_ids': body.brondocument_ids, 'aantal_brondocumenten': len(documents)},
            'generation_config': {'model': selected_model},
            'claude_model_used': selected_model,
            'claude_tokens_used': tokens_used,
            'created_by': get_user_id_from_request(request),
            'completed_at': datetime.utcnow().isoformat(),
            'created_at': datetime.utcnow().isoformat(),
        }

        save_result = db.table('ai_documents').insert(doc_data).execute()
        saved = save_result.data[0] if save_result.data else doc_data

        return {
            **saved,
            'document_content': generated_text,
            'claude_tokens_used': tokens_used,
            'heeft_downstream': body.template_key in DOWNSTREAM_TEMPLATES,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Generatie mislukt: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Generatie mislukt: {str(e)}")


# ============================================
# AKKOORD & DOWNSTREAM ENDPOINTS
# ============================================

@router.post("/documents/{document_id}/akkoord")
async def geef_akkoord(
    document_id: str,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        doc_result = db.table('ai_documents') \
            .select('*') \
            .eq('id', document_id) \
            .eq('is_deleted', False) \
            .single() \
            .execute()

        if not doc_result.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")

        doc = doc_result.data
        template_key = doc.get('template_key')
        tender_id = doc.get('tender_id')

        db.table('ai_documents').update({'status': 'completed', 'is_latest': True}).eq('id', document_id).execute()

        if template_key not in DOWNSTREAM_TEMPLATES:
            return {'success': True, 'document_id': document_id, 'heeft_downstream': False, 'message': 'Document goedgekeurd'}

        markdown = doc.get('document_content', '')
        parsed = parse_rode_draad_markdown(markdown)
        downstream_config = DOWNSTREAM_TEMPLATES[template_key]
        bestaande_data = {}

        if 'tenderplanning' in downstream_config['vult']:
            tp_result = db.table('milestones').select('id').eq('tender_id', tender_id).limit(1).execute()
            bestaande_data['tenderplanning'] = len(tp_result.data or []) > 0
        if 'projectplanning' in downstream_config['vult']:
            pp_result = db.table('planning_taken').select('id').eq('tender_id', tender_id).limit(1).execute()
            bestaande_data['projectplanning'] = len(pp_result.data or []) > 0
        if 'checklist' in downstream_config['vult']:
            cl_result = db.table('checklist_items').select('id').eq('tender_id', tender_id).limit(1).execute()
            bestaande_data['checklist'] = len(cl_result.data or []) > 0

        preview = {}
        for tab in downstream_config['vult']:
            if tab == 'tenderplanning':
                preview['tenderplanning'] = {'aantal': len(parsed['tenderplanning']), 'items': parsed['tenderplanning'][:3], 'heeft_bestaande_data': bestaande_data.get('tenderplanning', False), 'overschrijft': downstream_config['overschrijft']}
            elif tab == 'projectplanning':
                preview['projectplanning'] = {'aantal': len(parsed['projectplanning']), 'items': parsed['projectplanning'][:3], 'heeft_bestaande_data': bestaande_data.get('projectplanning', False), 'overschrijft': downstream_config['overschrijft']}
            elif tab == 'checklist':
                preview['checklist'] = {'aantal': len(parsed['checklist']), 'items': parsed['checklist'][:3], 'heeft_bestaande_data': bestaande_data.get('checklist', False), 'overschrijft': downstream_config['overschrijft']}
            elif tab == 'team':
                preview['team'] = {'aantal': len(parsed['team']), 'items': parsed['team'], 'heeft_bestaande_data': False, 'overschrijft': False}

        return {'success': True, 'document_id': document_id, 'heeft_downstream': True, 'overschrijft': downstream_config['overschrijft'], 'preview': preview, 'message': 'Document goedgekeurd. Bevestig welke tabs gevuld worden.'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Akkoord fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class DownstreamRequest(BaseModel):
    tabs: List[str]


@router.post("/documents/{document_id}/downstream")
async def voer_downstream_uit(
    document_id: str,
    body: DownstreamRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        doc_result = db.table('ai_documents').select('*').eq('id', document_id).single().execute()
        if not doc_result.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")

        doc = doc_result.data
        template_key = doc.get('template_key')
        tender_id = doc.get('tender_id')
        tenderbureau_id = doc.get('tenderbureau_id')
        user_id = get_user_id_from_request(request)
        markdown = doc.get('document_content', '')

        if template_key not in DOWNSTREAM_TEMPLATES:
            raise HTTPException(status_code=400, detail=f"Template '{template_key}' heeft geen downstream effect")

        downstream_config = DOWNSTREAM_TEMPLATES[template_key]
        mag_overschrijven = downstream_config['overschrijft']
        parsed = parse_rode_draad_markdown(markdown)
        resultaten = {}

        if 'tenderplanning' in body.tabs and parsed['tenderplanning']:
            bestaand = db.table('milestones').select('id').eq('tender_id', tender_id).limit(1).execute()
            heeft_data = len(bestaand.data or []) > 0
            if not heeft_data or mag_overschrijven:
                if mag_overschrijven and heeft_data:
                    db.table('milestones').delete().eq('tender_id', tender_id).execute()
                nieuwe_milestones = []
                for item in parsed['tenderplanning']:
                    mijlpaal_naam = (item.get('mijlpaal') or '').strip()
                    if not mijlpaal_naam or mijlpaal_naam.isdigit():
                        mijlpaal_naam = item.get('datum_tekst') or f"Mijlpaal {mijlpaal_naam or '?'}"
                    nieuwe_milestones.append({
                        'tender_id': tender_id,
                        'naam': mijlpaal_naam[:200],
                        'milestone_type': 'extern',
                        'datum': item['datum'] or datetime.now().date().isoformat(),
                        'status': 'pending',
                        'notities': item['datum_tekst'] if not item['datum'] else None,
                    })
                if nieuwe_milestones:
                    db.table('milestones').insert(nieuwe_milestones).execute()
                    resultaten['tenderplanning'] = {'aangemaakt': len(nieuwe_milestones), 'status': 'gevuld'}
            else:
                resultaten['tenderplanning'] = {'status': 'overgeslagen', 'reden': 'Bestaande data behouden'}

        if 'projectplanning' in body.tabs and parsed['projectplanning']:
            bestaand = db.table('planning_taken').select('id').eq('tender_id', tender_id).limit(1).execute()
            heeft_data = len(bestaand.data or []) > 0
            if not heeft_data or mag_overschrijven:
                if mag_overschrijven and heeft_data:
                    db.table('planning_taken').delete().eq('tender_id', tender_id).execute()
                nieuwe_taken = []
                for i, item in enumerate(parsed['projectplanning']):
                    taak = {
                        'tender_id': tender_id,
                        'tenderbureau_id': tenderbureau_id,
                        'taak_naam': item['taak_naam'],
                        'categorie': 'Projectplanning',
                        'status': 'todo',
                        'volgorde': i,
                        'created_by': user_id,
                    }
                    if item.get('deadline'):
                        taak['datum'] = f"{item['deadline']}T00:00:00+00:00"
                    if item.get('verantwoordelijke'):
                        taak['beschrijving'] = f"Verantwoordelijke: {item['verantwoordelijke']}"
                    nieuwe_taken.append(taak)
                if nieuwe_taken:
                    db.table('planning_taken').insert(nieuwe_taken).execute()
                    resultaten['projectplanning'] = {'aangemaakt': len(nieuwe_taken), 'status': 'gevuld'}
            else:
                resultaten['projectplanning'] = {'status': 'overgeslagen', 'reden': 'Bestaande data behouden'}

        if 'checklist' in body.tabs and parsed['checklist']:
            bestaand = db.table('checklist_items').select('id').eq('tender_id', tender_id).limit(1).execute()
            heeft_data = len(bestaand.data or []) > 0
            if not heeft_data or mag_overschrijven:
                if mag_overschrijven and heeft_data:
                    db.table('checklist_items').delete().eq('tender_id', tender_id).execute()
                nieuwe_items = []
                for i, item in enumerate(parsed['checklist']):
                    nieuwe_items.append({
                        'tender_id': tender_id,
                        'tenderbureau_id': tenderbureau_id,
                        'sectie': item.get('sectie', 'Inleverdocumenten'),
                        'taak_naam': item['taak_naam'],
                        'is_verplicht': item.get('is_verplicht', True),
                        'status': 'pending',
                        'volgorde': i,
                    })
                if nieuwe_items:
                    db.table('checklist_items').insert(nieuwe_items).execute()
                    resultaten['checklist'] = {'aangemaakt': len(nieuwe_items), 'status': 'gevuld'}
            else:
                resultaten['checklist'] = {'status': 'overgeslagen', 'reden': 'Bestaande data behouden'}

        if 'team' in body.tabs and parsed['team']:
            resultaten['team'] = {'status': 'suggesties', 'suggesties': parsed['team'], 'reden': 'Team suggesties — handmatig bevestigen vereist'}

        return {'success': True, 'document_id': document_id, 'tender_id': tender_id, 'resultaten': resultaten, 'message': f'Downstream uitgevoerd voor {len(resultaten)} tabs'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Downstream fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/documents/{document_id}/download-docx")
async def download_document_as_docx(
    document_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        doc_result = db.table('ai_documents') \
            .select('*, tenders(naam, opdrachtgever)') \
            .eq('id', document_id) \
            .eq('is_deleted', False) \
            .single() \
            .execute()

        if not doc_result.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")

        doc = doc_result.data
        markdown = doc.get('document_content', '')
        if not markdown:
            raise HTTPException(status_code=400, detail="Document heeft geen inhoud om te exporteren")

        template_key = doc.get('template_key', '')
        template_namen = {
            'offerte': 'Offerte Onderzoek',
            'rode_draad': 'Rode Draad Sessie',
            'versie1_inschrijving': 'Versie 1 Concept Inschrijving',
            'win_check': 'Tender Audit & Optimalisatie',
            'samenvatting': 'Tender Samenvatting',
            'checklist_extractor': 'Checklist Extractor',
        }
        template_naam = template_namen.get(template_key, template_key.replace('_', ' ').title())
        tender_data = doc.get('tenders') or {}
        tender_naam = tender_data.get('naam', '')
        opdrachtgever = tender_data.get('opdrachtgever', '')
        completed_at = doc.get('completed_at') or doc.get('created_at', '')
        try:
            dt = datetime.fromisoformat(completed_at.replace('Z', '+00:00'))
            gegenereerd_op = dt.strftime('%d-%m-%Y %H:%M')
        except Exception:
            gegenereerd_op = ''

        docx_bytes = convert_markdown_to_docx(
            markdown=markdown,
            tender_naam=tender_naam,
            template_naam=template_naam,
            opdrachtgever=opdrachtgever,
            gegenereerd_op=gegenereerd_op,
        )

        safe_tender = re.sub(r'[^\w\s-]', '', tender_naam)[:40].strip().replace(' ', '_')
        safe_template = template_key.replace('_', '-')
        datum_str = datetime.now().strftime('%Y%m%d')
        filename = f"TenderZen_{safe_template}_{safe_tender}_{datum_str}.docx"

        return StreamingResponse(
            io.BytesIO(docx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(docx_bytes)),
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ DOCX download fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Download mislukt: {str(e)}")


# ============================================
# v3.3: READ ENDPOINTS VOOR DOWNSTREAM DATA
# ============================================

@router.get("/tenders/{tender_id}/milestones")
async def get_tender_milestones(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('milestones') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .order('datum', desc=False) \
            .execute()
        items = result.data or []
        return {'success': True, 'items': items, 'total': len(items)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# v3.6: MILESTONE CRUD ENDPOINTS
# ============================================

class MilestoneUpdateRequest(BaseModel):
    naam: Optional[str] = None
    datum: Optional[str] = None
    tijd: Optional[str] = None
    status: Optional[str] = None
    milestone_type: Optional[str] = None
    notities: Optional[str] = None
    verantwoordelijke: Optional[str] = None


class MilestoneCreateRequest(BaseModel):
    naam: str
    milestone_type: str
    datum: Optional[str] = None
    tijd: Optional[str] = None
    status: str = 'pending'
    notities: Optional[str] = None


@router.patch("/tenders/{tender_id}/milestones/{milestone_id}")
async def update_milestone(
    tender_id: str,
    milestone_id: str,
    body: MilestoneUpdateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    """Update een bestaande milestone (datum, tijd, status, verantwoordelijke)."""
    try:
        bestaand = db.table('milestones') \
            .select('id, tender_id') \
            .eq('id', milestone_id) \
            .eq('tender_id', tender_id) \
            .single() \
            .execute()

        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Milestone niet gevonden")

        update_data = {}
        if body.naam is not None:
            update_data['naam'] = body.naam[:200]
        if body.datum is not None:
            update_data['datum'] = None if body.datum in ('', 'null') else body.datum
        if body.tijd is not None:
            update_data['tijd'] = body.tijd if body.tijd else None
        if body.status is not None:
            if body.status not in ('pending', 'completed', 'overdue'):
                raise HTTPException(status_code=400, detail="Status moet pending, completed of overdue zijn")
            update_data['status'] = body.status
        if body.milestone_type is not None:
            update_data['milestone_type'] = body.milestone_type
        if body.notities is not None:
            update_data['notities'] = body.notities[:500] if body.notities else None
        if body.verantwoordelijke is not None:
            update_data['verantwoordelijke'] = body.verantwoordelijke if body.verantwoordelijke else None

        if not update_data:
            raise HTTPException(status_code=400, detail="Geen velden om te updaten")

        result = db.table('milestones') \
            .update(update_data) \
            .eq('id', milestone_id) \
            .eq('tender_id', tender_id) \
            .execute()

        if not result.data:
            raise HTTPException(status_code=500, detail="Update mislukt")

        return {'success': True, 'milestone': result.data[0], 'message': 'Milestone bijgewerkt'}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Update milestone fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenders/{tender_id}/milestones", status_code=201)
async def create_milestone(
    tender_id: str,
    body: MilestoneCreateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    """Maak een nieuwe milestone aan voor een tender."""
    try:
        tender = db.table('tenders').select('id').eq('id', tender_id).single().execute()
        if not tender.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")

        milestone_data = {
            'tender_id': tender_id,
            'naam': body.naam[:200],
            'milestone_type': body.milestone_type,
            'datum': body.datum if body.datum else None,
            'tijd': body.tijd if body.tijd else None,
            'status': body.status,
            'notities': body.notities[:500] if body.notities else None,
        }

        result = db.table('milestones').insert(milestone_data).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Aanmaken mislukt")

        return {'success': True, 'milestone': result.data[0], 'message': 'Milestone aangemaakt'}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Create milestone fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/tenders/{tender_id}/milestones/{milestone_id}", status_code=200)
async def delete_milestone(
    tender_id: str,
    milestone_id: str,
    db: Client = Depends(get_supabase_async)
):
    """Verwijder een milestone."""
    try:
        bestaand = db.table('milestones') \
            .select('id') \
            .eq('id', milestone_id) \
            .eq('tender_id', tender_id) \
            .single() \
            .execute()

        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Milestone niet gevonden")

        db.table('milestones') \
            .delete() \
            .eq('id', milestone_id) \
            .eq('tender_id', tender_id) \
            .execute()

        return {'success': True, 'message': 'Milestone verwijderd'}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Delete milestone fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tenders/{tender_id}/planning-taken")
async def get_tender_planning_taken(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('planning_taken') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .order('volgorde', desc=False) \
            .execute()
        items = result.data or []
        return {'success': True, 'items': items, 'total': len(items)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tenders/{tender_id}/checklist-items")
async def get_tender_checklist_items(
    tender_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        result = db.table('checklist_items') \
            .select('*') \
            .eq('tender_id', tender_id) \
            .order('volgorde', desc=False) \
            .execute()
        items = result.data or []
        mapped = []
        for item in items:
            mapped.append({
                'id': item.get('id'),
                'tender_id': item.get('tender_id'),
                'naam': item.get('taak_naam') or item.get('naam') or '',
                'taak_naam': item.get('taak_naam') or '',
                'categorie': item.get('sectie') or item.get('categorie') or 'Overig',
                'sectie': item.get('sectie') or 'Overig',
                'status': item.get('status') or 'pending',
                'is_verplicht': item.get('is_verplicht', True),
                'verplicht': item.get('is_verplicht', True),
                'checked': item.get('status') == 'done',
                'volgorde': item.get('volgorde', 0),
            })
        done_count = len([i for i in mapped if i['status'] == 'done'])
        total = len(mapped)
        return {'success': True, 'items': mapped, 'total': total, 'done': done_count, 'badge': f"{done_count}/{total}" if total else ''}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# v3.4: AI EXTRACTIE + BACKPLANNING ENDPOINTS
# ============================================

PROMPT_EXTRACT_PLANNING = """
Je bent een tender-expert. Analyseer het aanbestedingsdocument en extraheer
alle externe planningsmijlpalen: publicatiedata, nota van inlichtingen,
indiendeadlines, gunningsdatum en contractstart.

Geef de output als JSON-array (geen markdown, geen uitleg):
[
  {
    "naam": "Publicatie aanbesteding",
    "datum": "2026-03-01",
    "datum_tekst": "1 maart 2026",
    "milestone_type": "extern",
    "notities": null
  }
]

Regels:
- datum altijd in ISO-formaat YYYY-MM-DD, of null als onbekend
- datum_tekst is de originele tekst uit het document
- milestone_type altijd "extern"
- Maximaal 15 mijlpalen
- Alleen daadwerkelijk genoemde data, geen aannames
"""

PROMPT_EXTRACT_CHECKLIST = """
Je bent een tender-expert. Analyseer het aanbestedingsdocument en extraheer
alle in te leveren documenten en eisen voor de inschrijving.

Geef de output als JSON-array (geen markdown, geen uitleg):
[
  {
    "taak_naam": "Inschrijfbiljet",
    "sectie": "Inleverdocumenten",
    "is_verplicht": true
  }
]

Regels:
- sectie is één van: "Inleverdocumenten", "Geschiktheidseisen",
  "Uitsluitingsgronden", "Gunningscriteria", "Overig"
- is_verplicht is true tenzij expliciet als optioneel beschreven
- taak_naam is kort en actiegericht (max 80 tekens)
- Maximaal 30 items
- Geen dubbelen
"""

PROMPT_GENERATE_BACKPLANNING = """
Je bent een tender-planningsexpert. Genereer een realistische interne
projectplanning op basis van de verstrekte gegevens.

Deadline indiening: {deadline}
Beschikbare teamleden en rollen: {team_info}
Vandaag: {vandaag}

Maak een terugwaartse planning met taken die op tijd afgerond kunnen worden
vóór de deadline. Verdeel taken logisch over 3 fasen:
- "Voorbereiding" — kickoff, analyse, structuur (eerste ~30% van de tijd)
- "Uitwerking" — schrijven, berekenen, reviewen (middelste ~50%)
- "Afronding & Indiening" — opmaak, eindcheck, indienen (laatste ~20%)

Geef de output als JSON-array (geen markdown, geen uitleg):
[
  {{
    "taak_naam": "Kickoff intern",
    "categorie": "Voorbereiding",
    "rol": "tendermanager",
    "datum": "2026-03-10",
    "beschrijving": "Opstartmeeting met het team",
    "volgorde": 1
  }}
]

Regels:
- datum in ISO-formaat YYYY-MM-DD
- categorie is ALTIJD één van: "Voorbereiding", "Uitwerking", "Afronding & Indiening"
- rol is één van: tendermanager, schrijver, calculator, reviewer, designer
- volgorde begint bij 1, oplopend
- Maximaal 20 taken
- Taken na vandaag en vóór of op de deadline
"""


async def _fetch_brondocumenten_voor_tender(
    tender_id: str,
    db: Client,
    max_docs: int = 3
) -> list[dict]:
    storage_items = []

    docs_result = db.table('tender_documents') \
        .select('original_file_name, file_name, storage_path') \
        .eq('tender_id', tender_id) \
        .eq('is_deleted', False) \
        .order('created_at', desc=False) \
        .execute()

    for doc in (docs_result.data or []):
        sp = doc.get('storage_path') or ''
        fn = doc.get('original_file_name') or doc.get('file_name') or 'document'
        if sp:
            storage_items.append({'storage_path': sp, 'filename': fn})

    if len(storage_items) < max_docs:
        try:
            tender_result = db.table('tenders').select('smart_import_id').eq('id', tender_id).single().execute()
            smart_import_id = (tender_result.data or {}).get('smart_import_id')
            if smart_import_id:
                si_result = db.table('smart_imports').select('uploaded_files, status').eq('id', smart_import_id).single().execute()
                si_data = si_result.data or {}
                if si_data.get('status') == 'completed':
                    for f in (si_data.get('uploaded_files') or []):
                        sp = f.get('storage_path') or ''
                        fn = f.get('original_name') or f.get('name') or 'document'
                        if sp:
                            storage_items.append({'storage_path': sp, 'filename': fn})
        except Exception as e:
            print(f"⚠️ Smart Import lookup mislukt: {e}")

    if not storage_items:
        return []

    content_blocks = []
    for item in storage_items[:max_docs]:
        storage_path = item['storage_path']
        filename = item['filename']
        file_bytes = fetch_document_from_storage(db, storage_path)
        if not file_bytes:
            continue
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext == 'pdf' and len(file_bytes) <= MAX_PDF_DIRECT_SIZE:
            content_blocks.append(prepare_pdf_for_claude(file_bytes, filename))
        elif ext in ('docx', 'doc'):
            tekst = extract_word_text(file_bytes)
            if tekst:
                content_blocks.append({'type': 'text', 'text': f"=== {filename} ===\n{tekst}"})
        elif ext == 'pdf':
            tekst = extract_pdf_text_fallback(file_bytes)
            if tekst:
                content_blocks.append({'type': 'text', 'text': f"=== {filename} (tekst-extractie) ===\n{tekst}"})

    return content_blocks


def _parse_json_response(raw: str) -> list:
    clean = raw.strip()
    clean = re.sub(r'^```(?:json)?\s*', '', clean)
    clean = re.sub(r'\s*```$', '', clean)
    clean = clean.strip()
    try:
        result = json.loads(clean)
        if isinstance(result, list):
            return result
        if isinstance(result, dict) and 'items' in result:
            return result['items']
        return []
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON parse fout: {e}\nRaw: {raw[:500]}")
        return []


# ── v3.6: Model selecteerbaar vanuit UI ──────────────────────────────────────
GELDIGE_EXTRACTIE_MODELLEN = {
    "claude-haiku-4-5-20251001",
    "claude-sonnet-4-6",
    "claude-opus-4-6",
}


class ExtractPlanningRequest(BaseModel):
    overschrijf: bool = False
    model: str = "claude-haiku-4-5-20251001"  # Selecteerbaar vanuit UI (Haiku/Sonnet/Opus)
# ─────────────────────────────────────────────────────────────────────────────


@router.post("/tenders/{tender_id}/extract-planning")
async def extract_planning(
    tender_id: str,
    body: ExtractPlanningRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        user_id = get_user_id_from_request(request)
        bestaand = db.table('milestones').select('id').eq('tender_id', tender_id).limit(1).execute()
        heeft_data = len(bestaand.data or []) > 0

        if heeft_data and not body.overschrijf:
            return {
                'success': False,
                'overgeslagen': True,
                'reden': 'Er zijn al milestones voor deze tender. Gebruik overschrijf=true om te vervangen.',
                'aangemaakt': 0
            }

        content_blocks = await _fetch_brondocumenten_voor_tender(tender_id, db, max_docs=3)
        if not content_blocks:
            raise HTTPException(status_code=422, detail="Geen brondocumenten gevonden voor deze tender.")

        # Haal de actieve prompt op uit de database (fallback naar hardcoded)
        prompt_tekst = PROMPT_EXTRACT_PLANNING
        try:
            prompt_result = db.table('ai_prompts') \
                .select('prompt_content') \
                .eq('template_key', 'planning_extractor') \
                .eq('status', 'active') \
                .is_('tenderbureau_id', 'null') \
                .order('version', desc=True) \
                .limit(1) \
                .execute()
            if prompt_result.data:
                prompt_tekst = prompt_result.data[0]['prompt_content']
                print(f"✅ Gebruik DB-prompt voor planning_extractor")
            else:
                print(f"⚠️ Geen actieve DB-prompt gevonden, gebruik hardcoded fallback")
        except Exception as prompt_err:
            print(f"⚠️ Prompt ophalen mislukt: {prompt_err} — gebruik hardcoded fallback")

        # Gebruik het gekozen model (valideer eerst)
        gekozen_model = body.model if body.model in GELDIGE_EXTRACTIE_MODELLEN else "claude-haiku-4-5-20251001"
        print(f"🤖 Extractie met model: {gekozen_model}")

        client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        response = client.messages.create(
            model=gekozen_model,
            max_tokens=2000,
            messages=[{
                'role': 'user',
                'content': content_blocks + [{'type': 'text', 'text': prompt_tekst}]
            }]
        )

        raw = response.content[0].text if response.content else ''
        items = _parse_json_response(raw)
        if not items:
            raise HTTPException(status_code=500, detail="Claude kon geen planningsmijlpalen extraheren.")

        if heeft_data and body.overschrijf:
            db.table('milestones').delete().eq('tender_id', tender_id).execute()

        GELDIGE_TYPES = {
            'publicatie', 'schouw', 'vragen_ronde_1', 'nota_inlichtingen_1',
            'vragen_ronde_2', 'nota_inlichtingen_2', 'vragen_ronde_3',
            'nota_inlichtingen_3', 'interne_deadline', 'sluitingsdatum',
            'alcatraz', 'presentatie', 'voorlopige_gunning', 'definitieve_gunning',
            'start_opdracht', 'einde_contract', 'overig'
        }

        nieuwe_milestones = []
        for item in items:
            if not item.get('naam'):
                continue
            milestone_type = item.get('milestone_type', 'overig')
            if milestone_type not in GELDIGE_TYPES:
                milestone_type = 'overig'
            nieuwe_milestones.append({
                'tender_id': tender_id,
                'naam': str(item['naam'])[:200],
                'milestone_type': milestone_type,
                'datum': item.get('datum') or None,
                'tijd': item.get('tijd') or None,
                'status': 'pending',
                'notities': item.get('notities') or None,
            })

        if nieuwe_milestones:
            db.table('milestones').insert(nieuwe_milestones).execute()

        # Sync milestones naar tender kolommen
        tender_service = TenderService(db)
        await tender_service.sync_milestones_to_tender(str(tender_id))

        return {
            'success': True,
            'aangemaakt': len(nieuwe_milestones),
            'items': nieuwe_milestones,
            'model_gebruikt': gekozen_model,
            'message': f'{len(nieuwe_milestones)} mijlpalen geëxtraheerd en opgeslagen'
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Extract planning fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenders/{tender_id}/extract-checklist")
async def extract_checklist(
    tender_id: str,
    body: ExtractPlanningRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        user_id = get_user_id_from_request(request)
        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        tenderbureau_id = tender_result.data.get('tenderbureau_id') if tender_result.data else None

        bestaand = db.table('checklist_items').select('id').eq('tender_id', tender_id).limit(1).execute()
        heeft_data = len(bestaand.data or []) > 0

        if heeft_data and not body.overschrijf:
            return {
                'success': False,
                'overgeslagen': True,
                'reden': 'Er is al een checklist voor deze tender. Gebruik overschrijf=true om te vervangen.',
                'aangemaakt': 0
            }

        content_blocks = await _fetch_brondocumenten_voor_tender(tender_id, db, max_docs=3)
        if not content_blocks:
            raise HTTPException(status_code=422, detail="Geen brondocumenten gevonden voor deze tender.")

        gekozen_model = body.model if body.model in GELDIGE_EXTRACTIE_MODELLEN else "claude-haiku-4-5-20251001"

        client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        response = client.messages.create(
            model=gekozen_model,
            max_tokens=2000,
            messages=[{
                'role': 'user',
                'content': content_blocks + [{'type': 'text', 'text': PROMPT_EXTRACT_CHECKLIST}]
            }]
        )

        raw = response.content[0].text if response.content else ''
        items = _parse_json_response(raw)
        if not items:
            raise HTTPException(status_code=500, detail="Claude kon geen checklist-items extraheren.")

        if heeft_data and body.overschrijf:
            db.table('checklist_items').delete().eq('tender_id', tender_id).execute()

        nieuwe_items = []
        for i, item in enumerate(items):
            if not item.get('taak_naam'):
                continue
            nieuwe_items.append({
                'tender_id': tender_id,
                'tenderbureau_id': tenderbureau_id,
                'taak_naam': str(item['taak_naam'])[:200],
                'sectie': item.get('sectie', 'Inleverdocumenten'),
                'is_verplicht': bool(item.get('is_verplicht', True)),
                'status': 'pending',
                'volgorde': i,
            })

        if nieuwe_items:
            db.table('checklist_items').insert(nieuwe_items).execute()

        return {
            'success': True,
            'aangemaakt': len(nieuwe_items),
            'items': nieuwe_items,
            'badge': f"0/{len(nieuwe_items)}",
            'message': f'{len(nieuwe_items)} checklist-items geëxtraheerd en opgeslagen'
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Extract checklist fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class BackplanningRequest(BaseModel):
    deadline: str
    overschrijf: bool = False
    team_assignments: Optional[dict] = None


@router.post("/tenders/{tender_id}/generate-backplanning")
async def generate_backplanning(
    tender_id: str,
    body: BackplanningRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        user_id = get_user_id_from_request(request)
        tender_result = db.table('tenders').select('tenderbureau_id, naam').eq('id', tender_id).single().execute()
        tender = tender_result.data or {}
        tenderbureau_id = tender.get('tenderbureau_id')

        bestaand = db.table('planning_taken').select('id').eq('tender_id', tender_id).limit(1).execute()
        heeft_data = len(bestaand.data or []) > 0

        if heeft_data and not body.overschrijf:
            return {
                'success': False,
                'overgeslagen': True,
                'reden': 'Er is al een projectplanning voor deze tender. Gebruik overschrijf=true om te vervangen.',
                'aangemaakt': 0
            }

        if body.team_assignments:
            team_info = '\n'.join(f"- {rol}: {naam}" for rol, naam in body.team_assignments.items())
        else:
            team_result = db.table('tender_team_assignments').select('rol_in_tender, user_id').eq('tender_id', tender_id).execute()
            team_rows = team_result.data or []
            if team_rows:
                user_ids = [r['user_id'] for r in team_rows if r.get('user_id')]
                namen_map = {}
                if user_ids:
                    leden_result = db.table('team_members').select('user_id, naam').in_('user_id', user_ids).execute()
                    namen_map = {r['user_id']: r.get('naam', 'onbekend') for r in (leden_result.data or [])}
                team_info = '\n'.join(
                    f"- {r.get('rol_in_tender', 'onbekend')}: {namen_map.get(r.get('user_id'), 'onbekend')}"
                    for r in team_rows
                )
            else:
                team_info = "Geen teamleden toegewezen — gebruik generieke rollen"

        vandaag = datetime.now().date().isoformat()
        prompt = PROMPT_GENERATE_BACKPLANNING.format(
            deadline=body.deadline,
            team_info=team_info,
            vandaag=vandaag
        )

        client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            messages=[{'role': 'user', 'content': prompt}]
        )

        raw = response.content[0].text if response.content else ''
        items = _parse_json_response(raw)
        if not items:
            raise HTTPException(status_code=500, detail="Claude kon geen projectplanning genereren.")

        if heeft_data and body.overschrijf:
            db.table('planning_taken').delete().eq('tender_id', tender_id).execute()

        nieuwe_taken = []
        for i, item in enumerate(items):
            if not item.get('taak_naam'):
                continue
            datum_str = item.get('datum')
            datum_iso = None
            if datum_str:
                parsed_datum = _parse_date(datum_str)
                if parsed_datum:
                    datum_iso = f"{parsed_datum}T00:00:00+00:00"

            taak = {
                'tender_id': tender_id,
                'tenderbureau_id': tenderbureau_id,
                'taak_naam': str(item['taak_naam'])[:200],
                'categorie': item.get('categorie') or 'Projectplanning',
                'status': 'todo',
                'volgorde': item.get('volgorde', i + 1),
                'created_by': user_id,
            }
            if datum_iso:
                taak['datum'] = datum_iso
            if item.get('beschrijving'):
                taak['beschrijving'] = str(item['beschrijving'])[:500]
            if item.get('rol'):
                taak['rol'] = item['rol']
            nieuwe_taken.append(taak)

        if nieuwe_taken:
            db.table('planning_taken').insert(nieuwe_taken).execute()

        return {
            'success': True,
            'aangemaakt': len(nieuwe_taken),
            'items': nieuwe_taken,
            'deadline': body.deadline,
            'message': f'{len(nieuwe_taken)} taken gegenereerd voor de projectplanning'
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Generate backplanning fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# v3.5: WRITE ENDPOINTS VOOR PLANNING TAKEN
# ============================================

class PlanningTaakUpdateRequest(BaseModel):
    taak_naam: Optional[str] = None
    status: Optional[str] = None
    datum: Optional[str] = None
    toegewezen_aan: Optional[List[str]] = None
    beschrijving: Optional[str] = None
    categorie: Optional[str] = None


@router.patch("/tenders/{tender_id}/planning-taken/{taak_id}")
async def update_planning_taak(
    tender_id: str,
    taak_id: str,
    body: PlanningTaakUpdateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        bestaand = db.table('planning_taken').select('id, tender_id').eq('id', taak_id).eq('tender_id', tender_id).single().execute()
        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Taak niet gevonden")

        update_data = {}
        if body.taak_naam is not None:
            update_data['taak_naam'] = body.taak_naam[:200]
        if body.status is not None:
            if body.status not in ('todo', 'active', 'done'):
                raise HTTPException(status_code=400, detail="Status moet todo, active of done zijn")
            update_data['status'] = body.status
        if body.datum is not None:
            update_data['datum'] = None if body.datum in ('', 'null') else body.datum
        if body.toegewezen_aan is not None:
            update_data['toegewezen_aan'] = body.toegewezen_aan
        if body.beschrijving is not None:
            update_data['beschrijving'] = body.beschrijving[:500] if body.beschrijving else None
        if body.categorie is not None:
            update_data['categorie'] = body.categorie
        if not update_data:
            raise HTTPException(status_code=400, detail="Geen velden om te updaten")

        result = db.table('planning_taken').update(update_data).eq('id', taak_id).eq('tender_id', tender_id).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Update mislukt")

        return {'success': True, 'taak': result.data[0], 'message': 'Taak bijgewerkt'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Update planning taak fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/tenders/{tender_id}/planning-taken/{taak_id}", status_code=200)
async def delete_planning_taak(
    tender_id: str,
    taak_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        bestaand = db.table('planning_taken').select('id').eq('id', taak_id).eq('tender_id', tender_id).single().execute()
        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Taak niet gevonden")
        db.table('planning_taken').delete().eq('id', taak_id).eq('tender_id', tender_id).execute()
        return {'success': True, 'message': 'Taak verwijderd'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Delete planning taak fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class NieuweTaakRequest(BaseModel):
    taak_naam: str
    categorie: Optional[str] = 'Projectplanning'
    status: Optional[str] = 'todo'
    datum: Optional[str] = None
    toegewezen_aan: Optional[List[str]] = None
    beschrijving: Optional[str] = None
    volgorde: Optional[int] = None


@router.post("/tenders/{tender_id}/planning-taken", status_code=201)
async def create_planning_taak(
    tender_id: str,
    body: NieuweTaakRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        user_id = get_user_id_from_request(request)
        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")
        tenderbureau_id = tender_result.data.get('tenderbureau_id')

        volgorde = body.volgorde
        if volgorde is None:
            max_result = db.table('planning_taken').select('volgorde').eq('tender_id', tender_id).order('volgorde', desc=True).limit(1).execute()
            volgorde = ((max_result.data[0]['volgorde'] or 0) + 1) if max_result.data else 1

        taak_data = {
            'tender_id': tender_id,
            'tenderbureau_id': tenderbureau_id,
            'taak_naam': body.taak_naam[:200],
            'categorie': body.categorie or 'Projectplanning',
            'status': body.status or 'todo',
            'volgorde': volgorde,
            'created_by': user_id,
        }
        if body.datum:
            taak_data['datum'] = body.datum
        if body.toegewezen_aan is not None:
            taak_data['toegewezen_aan'] = body.toegewezen_aan
        if body.beschrijving:
            taak_data['beschrijving'] = body.beschrijving[:500]

        result = db.table('planning_taken').insert(taak_data).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Aanmaken mislukt")

        return {'success': True, 'taak': result.data[0], 'message': 'Taak aangemaakt'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Create planning taak fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _pp_fase_van_rol(rol: str) -> str:
    rol = (rol or '').lower()
    if rol in ('tendermanager',):
        return 'Voorbereiding'
    if rol in ('schrijver', 'calculator', 'reviewer', 'designer'):
        return 'Uitwerking'
    if rol in ('sales', 'directie'):
        return 'Afronding & Indiening'
    return 'Projectplanning'


class PopulateFromTemplateRequest(BaseModel):
    template_id: str
    overschrijf: bool = False


@router.post("/tenders/{tender_id}/populate-from-template")
async def populate_planning_from_template(
    tender_id: str,
    body: PopulateFromTemplateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        user_id = get_user_id_from_request(request)
        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")
        tenderbureau_id = tender_result.data.get('tenderbureau_id')

        template_result = db.table('planning_templates').select('id, naam').eq('id', body.template_id).single().execute()
        if not template_result.data:
            raise HTTPException(status_code=404, detail="Template niet gevonden")

        taken_result = db.table('planning_template_taken').select('*').eq('template_id', body.template_id).order('volgorde').execute()
        template_taken = taken_result.data or []
        if not template_taken:
            raise HTTPException(status_code=422, detail="Template heeft geen taken")

        bestaand = db.table('planning_taken').select('id').eq('tender_id', tender_id).limit(1).execute()
        heeft_data = len(bestaand.data or []) > 0

        if heeft_data and not body.overschrijf:
            return {'success': False, 'overgeslagen': True, 'reden': 'Er is al een projectplanning. Gebruik overschrijf=true.', 'aangemaakt': 0}

        if heeft_data and body.overschrijf:
            db.table('planning_taken').delete().eq('tender_id', tender_id).execute()

        nieuwe_taken = []
        for i, tt in enumerate(template_taken):
            taak = {
                'tender_id': tender_id,
                'tenderbureau_id': tenderbureau_id,
                'taak_naam': tt.get('naam', 'Taak')[:200],
                'categorie': tt.get('categorie') or tt.get('fase') or _pp_fase_van_rol(tt.get('rol', '')),
                'status': 'todo',
                'volgorde': tt.get('volgorde', i),
                'created_by': user_id,
            }
            if tt.get('beschrijving'):
                taak['beschrijving'] = tt['beschrijving'][:500]
            if tt.get('rol'):
                taak['rol'] = tt['rol']
            nieuwe_taken.append(taak)

        result = db.table('planning_taken').insert(nieuwe_taken).execute()
        aangemaakt = len(result.data or [])

        return {'success': True, 'aangemaakt': aangemaakt, 'template_naam': template_result.data['naam'], 'items': result.data or [], 'message': f'{aangemaakt} taken geladen vanuit template'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Populate from template fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# v3.5: CHECKLIST WRITE ENDPOINTS
# ============================================

class ChecklistItemUpdateRequest(BaseModel):
    taak_naam: Optional[str] = None
    status: Optional[str] = None
    deadline: Optional[str] = None
    toegewezen_aan: Optional[List[str]] = None
    beschrijving: Optional[str] = None
    sectie: Optional[str] = None
    is_verplicht: Optional[bool] = None


class NieuwChecklistItemRequest(BaseModel):
    taak_naam: str
    sectie: Optional[str] = 'Overige documenten'
    is_verplicht: Optional[bool] = False
    status: Optional[str] = 'pending'
    volgorde: Optional[int] = None
    beschrijving: Optional[str] = None
    toegewezen_aan: Optional[List[str]] = None
    deadline: Optional[str] = None


class PopulateChecklistFromTemplateRequest(BaseModel):
    template_id: str
    overschrijf: bool = False


@router.patch("/tenders/{tender_id}/checklist-items/{item_id}")
async def update_checklist_item(
    tender_id: str,
    item_id: str,
    body: ChecklistItemUpdateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        bestaand = db.table('checklist_items').select('id, tender_id').eq('id', item_id).eq('tender_id', tender_id).single().execute()
        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Checklist item niet gevonden")

        update_data = {}
        if body.taak_naam is not None:
            update_data['taak_naam'] = body.taak_naam[:300]
        if body.status is not None:
            if body.status not in ('pending', 'in_progress', 'done'):
                raise HTTPException(status_code=400, detail="Status moet pending, in_progress of done zijn")
            update_data['status'] = body.status
        if body.deadline is not None:
            update_data['deadline'] = None if body.deadline in ('', 'null') else body.deadline
        if body.toegewezen_aan is not None:
            update_data['verantwoordelijke'] = body.toegewezen_aan[0] if body.toegewezen_aan else None
            try:
                update_data['toegewezen_aan'] = body.toegewezen_aan
            except Exception:
                pass
        if body.beschrijving is not None:
            update_data['beschrijving'] = body.beschrijving[:500] if body.beschrijving else None
        if body.sectie is not None:
            update_data['sectie'] = body.sectie
        if body.is_verplicht is not None:
            update_data['is_verplicht'] = body.is_verplicht
        if not update_data:
            raise HTTPException(status_code=400, detail="Geen velden om te updaten")

        result = db.table('checklist_items').update(update_data).eq('id', item_id).eq('tender_id', tender_id).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Update mislukt")

        return {'success': True, 'item': result.data[0], 'message': 'Item bijgewerkt'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Update checklist item fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/tenders/{tender_id}/checklist-items/{item_id}", status_code=200)
async def delete_checklist_item(
    tender_id: str,
    item_id: str,
    db: Client = Depends(get_supabase_async)
):
    try:
        bestaand = db.table('checklist_items').select('id').eq('id', item_id).eq('tender_id', tender_id).single().execute()
        if not bestaand.data:
            raise HTTPException(status_code=404, detail="Checklist item niet gevonden")
        db.table('checklist_items').delete().eq('id', item_id).eq('tender_id', tender_id).execute()
        return {'success': True, 'message': 'Item verwijderd'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Delete checklist item fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenders/{tender_id}/checklist-items", status_code=201)
async def create_checklist_item(
    tender_id: str,
    body: NieuwChecklistItemRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")
        tenderbureau_id = tender_result.data.get('tenderbureau_id')

        volgorde = body.volgorde
        if volgorde is None:
            max_result = db.table('checklist_items').select('volgorde').eq('tender_id', tender_id).order('volgorde', desc=True).limit(1).execute()
            volgorde = ((max_result.data[0]['volgorde'] or 0) + 1) if max_result.data else 1

        item_data = {
            'tender_id': tender_id,
            'tenderbureau_id': tenderbureau_id,
            'taak_naam': body.taak_naam[:300],
            'sectie': body.sectie or 'Overige documenten',
            'is_verplicht': body.is_verplicht if body.is_verplicht is not None else False,
            'status': body.status or 'pending',
            'volgorde': volgorde,
        }
        if body.beschrijving:
            item_data['beschrijving'] = body.beschrijving[:500]
        if body.deadline:
            item_data['deadline'] = body.deadline
        if body.toegewezen_aan is not None:
            item_data['verantwoordelijke'] = body.toegewezen_aan[0] if body.toegewezen_aan else None

        result = db.table('checklist_items').insert(item_data).execute()
        if not result.data:
            raise HTTPException(status_code=500, detail="Aanmaken mislukt")

        return {'success': True, 'item': result.data[0], 'message': 'Item aangemaakt'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Create checklist item fout: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenders/{tender_id}/populate-checklist-from-template")
async def populate_checklist_from_template(
    tender_id: str,
    body: PopulateChecklistFromTemplateRequest,
    request: Request,
    db: Client = Depends(get_supabase_async)
):
    try:
        tender_result = db.table('tenders').select('tenderbureau_id').eq('id', tender_id).single().execute()
        if not tender_result.data:
            raise HTTPException(status_code=404, detail="Tender niet gevonden")
        tenderbureau_id = tender_result.data.get('tenderbureau_id')

        template_result = db.table('planning_templates').select('id, naam').eq('id', body.template_id).single().execute()
        if not template_result.data:
            raise HTTPException(status_code=404, detail="Template niet gevonden")

        taken_result = db.table('planning_template_taken').select('*').eq('template_id', body.template_id).order('volgorde').execute()
        template_taken = taken_result.data or []
        if not template_taken:
            raise HTTPException(status_code=422, detail="Template heeft geen items")

        bestaand = db.table('checklist_items').select('id').eq('tender_id', tender_id).limit(1).execute()
        heeft_data = len(bestaand.data or []) > 0

        if heeft_data and not body.overschrijf:
            return {'success': False, 'overgeslagen': True, 'reden': 'Er zijn al checklist items. Gebruik overschrijf=true.', 'aangemaakt': 0}

        if heeft_data and body.overschrijf:
            db.table('checklist_items').delete().eq('tender_id', tender_id).execute()

        nieuwe_items = []
        for i, tt in enumerate(template_taken):
            item = {
                'tender_id': tender_id,
                'tenderbureau_id': tenderbureau_id,
                'taak_naam': tt.get('naam', 'Item')[:300],
                'sectie': tt.get('sectie') or tt.get('categorie') or tt.get('fase') or 'Overige documenten',
                'is_verplicht': bool(tt.get('is_verplicht', False)),
                'status': 'pending',
                'volgorde': tt.get('volgorde', i),
            }
            if tt.get('beschrijving'):
                item['beschrijving'] = tt['beschrijving'][:500]
            nieuwe_items.append(item)

        result = db.table('checklist_items').insert(nieuwe_items).execute()
        aangemaakt = len(result.data or [])

        return {'success': True, 'aangemaakt': aangemaakt, 'template_naam': template_result.data['naam'], 'items': result.data or [], 'message': f'{aangemaakt} items geladen vanuit template'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Populate checklist from template fout: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))