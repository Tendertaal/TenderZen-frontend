"""
Implementatieplanning API — TenderZen
CRUD voor implementatie secties/taken/metadata, AI-generatie en Excel/PDF export.
"""
import asyncio
import base64
import io
import json
import logging
from datetime import datetime, date, timezone, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from supabase import Client

from app.core.database import get_supabase_async
from app.core.dependencies import get_current_user
from app.services.anthropic_service import call_claude

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Implementatieplanning"])

CLAUDE_MODEL = "claude-haiku-4-5-20251001"
MAX_TOKENS   = 8000

# Exacte model-strings zoals gebruikt in ai_documents.py
TOEGESTANE_MODELLEN = {
    "claude-haiku-4-5-20251001",
    "claude-sonnet-4-6",
    "claude-opus-4-6",
}

SYSTEEM_PROMPT = """Je bent een expert in aanbestedingen en projectplanning. \
Genereer een implementatieplanning voor een inschrijving op een aanbesteding.

Gebruik maximaal 5 secties met elk maximaal 6 taken (totaal max 30 taken).
Houd toelichting per taak kort (max 1 zin).

Secties (kies de meest relevante):
1. Initiatiefase / Opstartfase
2. Voorbereiding / Werkvoorbereiding
3. Uitvoering
4. Oplevering & Nazorg
5. Beheer & Administratie

Per taak: nummer, naam, verantwoordelijke, toelichting, startdatum, einddatum, dagen.
Baseer datums op de opgegeven startdatum.
Retourneer UITSLUITEND een JSON object, geen uitleg, geen markdown backticks."""

CHAT_SYSTEEM_PROMPT = """Je bent een expert projectplanner. De gebruiker heeft een \
implementatieplanning gemaakt en wil die verfijnen via instructies in natuurlijke taal.

Je krijgt de volledige huidige planning als JSON mee.
Analyseer de instructie en geef een gestructureerd diff terug.

Retourneer UITSLUITEND dit JSON formaat, geen uitleg, geen markdown:
{
  "samenvatting": "Korte beschrijving van wat je hebt aangepast (max 1 zin)",
  "wijzigingen": [
    {
      "type": "sectie_toevoegen",
      "data": {
        "naam": "Evaluatiefase",
        "kleur": "#a7f3d0",
        "volgorde": 5,
        "taken": [
          {
            "nummer": "E.1",
            "naam": "Kwaliteitscontrole",
            "verantwoordelijke": "Opdrachtgever",
            "toelichting": "Eindcontrole per fase",
            "startdatum": "YYYY-MM-DD",
            "einddatum": "YYYY-MM-DD",
            "dagen": 10,
            "volgorde": 1
          }
        ]
      }
    },
    {
      "type": "taak_toevoegen",
      "sectie_id": "uuid-bestaande-sectie",
      "data": { "nummer": "X.1", "naam": "...", "startdatum": "YYYY-MM-DD", "einddatum": "YYYY-MM-DD", "dagen": 5, "volgorde": 99 }
    },
    {
      "type": "taak_wijzigen",
      "taak_id": "uuid",
      "data": { "startdatum": "YYYY-MM-DD", "einddatum": "YYYY-MM-DD" }
    },
    {
      "type": "taak_verwijderen",
      "taak_id": "uuid"
    },
    {
      "type": "sectie_wijzigen",
      "sectie_id": "uuid",
      "data": { "naam": "Nieuwe naam", "kleur": "#..." }
    },
    {
      "type": "sectie_verwijderen",
      "sectie_id": "uuid"
    }
  ]
}

Toegestane types: sectie_toevoegen, sectie_wijzigen, sectie_verwijderen,
                  taak_toevoegen, taak_wijzigen, taak_verwijderen.
Houd datums realistisch t.o.v. planstart/planeinde in de planning.
Gebruik YYYY-MM-DD formaat voor alle datums."""

VRAAG_SYSTEEM_PROMPT = """Je bent een expert projectplanner. Je krijgt een \
implementatieplanning als JSON en een vraag van de gebruiker.

Analyseer de planning en beantwoord de vraag op basis van de gegevens.
Wijs op relevante details zoals verantwoordelijken, datums, doorlooptijden,
overlappende fasen, of risico's die je ziet.

Retourneer UITSLUITEND dit JSON formaat, geen uitleg, geen markdown backticks:
{
  "antwoord": "Je volledige antwoord hier als leesbare tekst. Mag meerdere zinnen zijn."
}

Wijzig NOOIT de planning. Voer NOOIT acties uit. Alleen analyseren en beantwoorden."""

# ── Pydantic models ──────────────────────────────────────────────────────────

class MetadataUpdate(BaseModel):
    projectnaam:   Optional[str] = None
    opdrachtgever: Optional[str] = None
    opdrachtnemer: Optional[str] = None
    planstart:     Optional[str] = None
    planeinde:     Optional[str] = None
    notities:      Optional[str] = None

class SectieCreate(BaseModel):
    naam:     str
    kleur:    Optional[str] = "#c7d2fe"
    volgorde: int = 0

class SectieUpdate(BaseModel):
    naam:     Optional[str] = None
    kleur:    Optional[str] = None
    volgorde: Optional[int] = None

class TaakCreate(BaseModel):
    sectie_id:        str
    nummer:           Optional[str] = None
    naam:             str
    verantwoordelijke: Optional[str] = None
    toelichting:      Optional[str] = None
    status:           str = "open"
    startdatum:       Optional[str] = None
    einddatum:        Optional[str] = None
    dagen:            Optional[int] = None
    volgorde:         int = 0

class TaakUpdate(BaseModel):
    sectie_id:        Optional[str] = None
    nummer:           Optional[str] = None
    naam:             Optional[str] = None
    verantwoordelijke: Optional[str] = None
    toelichting:      Optional[str] = None
    status:           Optional[str] = None
    startdatum:       Optional[str] = None
    einddatum:        Optional[str] = None
    dagen:            Optional[int] = None
    volgorde:         Optional[int] = None

class Document(BaseModel):
    base64: str
    naam:   str

class GenereerRequest(BaseModel):
    documenten:          Optional[List[Document]] = []
    tender_omschrijving: Optional[str] = None
    projectnaam:         Optional[str] = None
    opdrachtgever:       Optional[str] = None
    opdrachtnemer:       Optional[str] = None
    planstart:           Optional[str] = None
    model:               str = "claude-haiku-4-5-20251001"

class ChatRequest(BaseModel):
    bericht:  str
    model:    str = "claude-sonnet-4-6"
    planning: dict  # { metadata: {...}, secties: [{naam, taken:[...]}] }
    modus:    str = "aanpas"  # "aanpas" of "vraag"

# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_tender(db: Client, tender_id: str) -> dict:
    res = db.table("tenders").select("id,tenderbureau_id,naam,omschrijving,opdrachtgever") \
        .eq("id", tender_id).execute()
    if not res.data:
        raise HTTPException(404, f"Tender {tender_id} niet gevonden")
    return res.data[0]

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()

def _detecteer_mime(naam: str) -> str:
    n = naam.lower()
    if n.endswith('.pdf'):  return "application/pdf"
    if n.endswith('.docx'): return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if n.endswith('.doc'):  return "application/msword"
    return "application/pdf"  # fallback

def _bereken_dagen(startdatum: Optional[str], einddatum: Optional[str]) -> Optional[int]:
    """Bereken aantal werkdagen (kalender) inclusief start- en einddatum."""
    if not startdatum or not einddatum:
        return None
    try:
        s = date.fromisoformat(startdatum)
        e = date.fromisoformat(einddatum)
        return max(1, (e - s).days + 1)
    except Exception:
        return None

def _load_planning(db: Client, tender_id: str) -> dict:
    """Laad metadata + secties + taken voor een tender."""
    meta_res = db.table("implementatie_metadata").select("*").eq("tender_id", tender_id).execute()
    metadata = meta_res.data[0] if meta_res.data else None

    secties_res = db.table("implementatie_secties") \
        .select("*").eq("tender_id", tender_id).order("volgorde").execute()
    secties = secties_res.data or []

    if secties:
        sectie_ids = [s["id"] for s in secties]
        taken_res  = db.table("implementatie_taken") \
            .select("*").in_("sectie_id", sectie_ids).order("volgorde").execute()
        taken_map  = {}
        for t in (taken_res.data or []):
            taken_map.setdefault(t["sectie_id"], []).append(t)
        for s in secties:
            s["taken"] = taken_map.get(s["id"], [])

    return {"metadata": metadata, "secties": secties}

# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

# ── GET planning ─────────────────────────────────────────────────────────────

@router.get("/implementatieplanning/{tender_id}")
async def get_planning(
    tender_id:    str,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    return _load_planning(db, tender_id)


# ── AI generatie ─────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/{tender_id}/genereer")
async def genereer_planning(
    tender_id:    str,
    body:         GenereerRequest,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    gekozen_model = body.model if body.model in TOEGESTANE_MODELLEN else CLAUDE_MODEL
    tender     = _require_tender(db, tender_id)
    bureau_id  = tender["tenderbureau_id"]

    projectnaam   = body.projectnaam   or tender.get("naam")          or "Project"
    opdrachtgever = body.opdrachtgever or tender.get("opdrachtgever") or "Opdrachtgever"
    opdrachtnemer = body.opdrachtnemer or "Opdrachtnemer"
    planstart     = body.planstart     or date.today().isoformat()
    omschrijving  = (body.tender_omschrijving or tender.get("omschrijving") or "")[:3000]

    user_prompt = f"""Projectnaam: {projectnaam}
Opdrachtgever: {opdrachtgever}
Opdrachtnemer: {opdrachtnemer}
Startdatum: {planstart}
{f"Tenderomschrijving: {omschrijving}" if omschrijving else ""}

Genereer een implementatieplanning in exact dit JSON formaat:
{{
  "projectnaam": "{projectnaam}",
  "opdrachtgever": "{opdrachtgever}",
  "planstart": "{planstart}",
  "planeinde": "YYYY-MM-DD",
  "secties": [
    {{
      "naam": "Sectienaam",
      "kleur": "#c7d2fe",
      "volgorde": 1,
      "taken": [
        {{
          "nummer": "O.1",
          "naam": "Taaknaam",
          "verantwoordelijke": "Verantwoordelijke partij",
          "toelichting": "Korte toelichting",
          "startdatum": "YYYY-MM-DD",
          "einddatum": "YYYY-MM-DD",
          "dagen": 1,
          "volgorde": 1
        }}
      ]
    }}
  ]
}}"""

    # Valideer documentenaantal
    documenten = body.documenten or []
    if len(documenten) > 5:
        raise HTTPException(400, "Maximaal 5 documenten toegestaan")

    # Bouw messages — tekst-only als fallback, document-blocks als documenten aangeleverd
    if documenten:
        content: list = []
        for doc in documenten:
            try:
                content.append({
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": _detecteer_mime(doc.naam),
                        "data": doc.base64,
                    }
                })
            except Exception as e:
                logger.warning(f"[ip] Document '{doc.naam}' overgeslagen: {e}")
        content.append({"type": "text", "text": user_prompt})
        messages = [{"role": "user", "content": content}]
    else:
        messages = [{"role": "user", "content": user_prompt}]

    # Claude aanroepen
    try:
        resp = await asyncio.to_thread(
            call_claude,
            messages=messages,
            model=gekozen_model,
            max_tokens=MAX_TOKENS,
            system=SYSTEEM_PROMPT,
            log_usage=True,
            db=db,
            tender_id=tender_id,
            bureau_id=bureau_id,
            call_type="implementatieplanning",
        )
        raw        = resp.content[0].text.strip()
        stop_reason = getattr(resp, "stop_reason", None)

        # Detecteer afgekapte response (token-limiet bereikt)
        if stop_reason == "max_tokens":
            logger.warning(f"[ip] Claude response afgekapt (max_tokens). Ruwe output ({len(raw)} tekens).")
            raise ValueError(
                f"Claude response afgekapt — de planning is te groot voor het token-limiet. "
                f"Probeer een kortere omschrijving, of verwijder het document."
            )

        start = raw.find("{")
        eind  = raw.rfind("}") + 1
        if start == -1 or eind == 0:
            raise ValueError("Geen JSON in Claude output")

        json_str = raw[start:eind]
        try:
            ai_data = json.loads(json_str)
        except json.JSONDecodeError as json_err:
            # JSON is alsnog kapot (bijv. door onvolledige output) — log fragment voor debugging
            frag = json_str[-200:] if len(json_str) > 200 else json_str
            logger.error(f"[ip] JSON parse fout: {json_err}. Laatste 200 tekens: {frag!r}")
            raise ValueError(f"Claude gaf ongeldige JSON terug: {json_err}")
    except Exception as e:
        logger.error(f"[ip] Claude fout: {e}")
        raise HTTPException(500, f"AI-generatie mislukt: {str(e)}")

    # Bestaande data verwijderen
    db.table("implementatie_secties").delete().eq("tender_id", tender_id).execute()

    # Metadata opslaan
    meta_row = {
        "tender_id":        tender_id,
        "tenderbureau_id":  bureau_id,
        "projectnaam":      ai_data.get("projectnaam",   projectnaam),
        "opdrachtgever":    ai_data.get("opdrachtgever", opdrachtgever),
        "opdrachtnemer":    opdrachtnemer,
        "planstart":        ai_data.get("planstart",     planstart),
        "planeinde":        ai_data.get("planeinde"),
        "ai_gegenereerd":   True,
        "ai_gegenereerd_op": _now(),
        "updated_at":       _now(),
    }
    db.table("implementatie_metadata").upsert(meta_row, on_conflict="tender_id").execute()

    # Secties en taken
    totaal_taken = 0
    for sectie_data in ai_data.get("secties", []):
        sectie_res = db.table("implementatie_secties").insert({
            "tender_id":      tender_id,
            "tenderbureau_id": bureau_id,
            "naam":           sectie_data.get("naam",     "Fase"),
            "kleur":          sectie_data.get("kleur",    "#c7d2fe"),
            "volgorde":       sectie_data.get("volgorde", 0),
        }).execute()
        sectie_id = sectie_res.data[0]["id"]

        for taak_data in sectie_data.get("taken", []):
            db.table("implementatie_taken").insert({
                "sectie_id":       sectie_id,
                "tender_id":       tender_id,
                "tenderbureau_id": bureau_id,
                "nummer":          taak_data.get("nummer"),
                "naam":            taak_data.get("naam",    "Taak"),
                "verantwoordelijke": taak_data.get("verantwoordelijke"),
                "toelichting":     taak_data.get("toelichting"),
                "startdatum":      taak_data.get("startdatum"),
                "einddatum":       taak_data.get("einddatum"),
                "dagen":           taak_data.get("dagen"),
                "volgorde":        taak_data.get("volgorde", 0),
                "status":          "open",
            }).execute()
            totaal_taken += 1

    return {
        "succes": True,
        "totaal_taken": totaal_taken,
        "bericht": f"Planning gegenereerd met {totaal_taken} taken.",
    }


# ── Metadata ──────────────────────────────────────────────────────────────────

@router.put("/implementatieplanning/{tender_id}/metadata")
async def update_metadata(
    tender_id:    str,
    body:         MetadataUpdate,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    tender    = _require_tender(db, tender_id)
    bureau_id = tender["tenderbureau_id"]

    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    updates["updated_at"] = _now()

    existing = db.table("implementatie_metadata").select("id").eq("tender_id", tender_id).execute()
    if existing.data:
        db.table("implementatie_metadata").update(updates).eq("tender_id", tender_id).execute()
    else:
        updates["tender_id"]       = tender_id
        updates["tenderbureau_id"] = bureau_id
        db.table("implementatie_metadata").insert(updates).execute()

    return {"succes": True}


# ── Secties ───────────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/{tender_id}/secties")
async def add_sectie(
    tender_id:    str,
    body:         SectieCreate,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    tender = _require_tender(db, tender_id)

    # Volgorde: hoogste bestaande + 1
    vol_res  = db.table("implementatie_secties").select("volgorde") \
        .eq("tender_id", tender_id).order("volgorde", desc=True).limit(1).execute()
    volgorde = (vol_res.data[0]["volgorde"] + 1) if vol_res.data else 1

    res = db.table("implementatie_secties").insert({
        "tender_id":       tender_id,
        "tenderbureau_id": tender["tenderbureau_id"],
        "naam":            body.naam,
        "kleur":           body.kleur or "#c7d2fe",
        "volgorde":        volgorde,
    }).execute()
    return res.data[0]


@router.put("/implementatieplanning/secties/{sectie_id}")
async def update_sectie(
    sectie_id:    str,
    body:         SectieUpdate,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    updates["updated_at"] = _now()
    db.table("implementatie_secties").update(updates).eq("id", sectie_id).execute()
    return {"succes": True}


@router.delete("/implementatieplanning/secties/{sectie_id}")
async def delete_sectie(
    sectie_id:    str,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    # taken worden gecascaded via FK, maar doen we ook expliciet voor zekerheid
    db.table("implementatie_taken").delete().eq("sectie_id", sectie_id).execute()
    db.table("implementatie_secties").delete().eq("id", sectie_id).execute()
    return {"succes": True}


# ── Taken ─────────────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/taken")
async def add_taak(
    body:         TaakCreate,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    sectie_res = db.table("implementatie_secties") \
        .select("tender_id,tenderbureau_id").eq("id", body.sectie_id).execute()
    if not sectie_res.data:
        raise HTTPException(404, "Sectie niet gevonden")
    sectie = sectie_res.data[0]

    dagen = body.dagen if body.dagen is not None \
        else _bereken_dagen(body.startdatum, body.einddatum)

    res = db.table("implementatie_taken").insert({
        "sectie_id":         body.sectie_id,
        "tender_id":         sectie["tender_id"],
        "tenderbureau_id":   sectie["tenderbureau_id"],
        "nummer":            body.nummer,
        "naam":              body.naam,
        "verantwoordelijke": body.verantwoordelijke,
        "toelichting":       body.toelichting,
        "status":            body.status,
        "startdatum":        body.startdatum,
        "einddatum":         body.einddatum,
        "dagen":             dagen,
        "volgorde":          body.volgorde,
    }).execute()
    return res.data[0]


@router.put("/implementatieplanning/taken/{taak_id}")
async def update_taak(
    taak_id:      str,
    body:         TaakUpdate,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}

    # Herbereken dagen als start- of einddatum gewijzigd
    if "startdatum" in updates or "einddatum" in updates:
        # Haal huidige waarden op om ontbrekende kant op te vullen
        if "startdatum" not in updates or "einddatum" not in updates:
            huidig = db.table("implementatie_taken").select("startdatum,einddatum") \
                .eq("id", taak_id).execute()
            if huidig.data:
                huidig_row = huidig.data[0]
                start = updates.get("startdatum") or huidig_row.get("startdatum")
                einde = updates.get("einddatum")  or huidig_row.get("einddatum")
            else:
                start = updates.get("startdatum")
                einde = updates.get("einddatum")
        else:
            start = updates["startdatum"]
            einde = updates["einddatum"]
        berekend = _bereken_dagen(start, einde)
        if berekend is not None:
            updates["dagen"] = berekend

    updates["updated_at"] = _now()
    db.table("implementatie_taken").update(updates).eq("id", taak_id).execute()
    return {"succes": True}


@router.delete("/implementatieplanning/taken/{taak_id}")
async def delete_taak(
    taak_id:      str,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    db.table("implementatie_taken").delete().eq("id", taak_id).execute()
    return {"succes": True}


# ── AI chat ───────────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/{tender_id}/chat")
async def chat_planning(
    tender_id:    str,
    body:         ChatRequest,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    if body.modus not in ("aanpas", "vraag"):
        raise HTTPException(400, "Ongeldige modus — gebruik 'aanpas' of 'vraag'")

    gekozen_model = body.model if body.model in TOEGESTANE_MODELLEN else "claude-sonnet-4-6"
    tender    = _require_tender(db, tender_id)
    bureau_id = tender["tenderbureau_id"]

    label_bericht = "Instructie" if body.modus == "aanpas" else "Vraag"
    user_message = (
        f"Huidige planning:\n{json.dumps(body.planning, ensure_ascii=False, indent=2)}\n\n"
        f"{label_bericht} van de gebruiker:\n{body.bericht}"
    )

    # ── Vraag-modus: alleen analyseren, niets toepassen ──────────────────────
    if body.modus == "vraag":
        try:
            resp = await asyncio.to_thread(
                call_claude,
                messages=[{"role": "user", "content": user_message}],
                model=gekozen_model,
                max_tokens=1000,
                system=VRAAG_SYSTEEM_PROMPT,
                log_usage=True,
                db=db,
                tender_id=tender_id,
                bureau_id=bureau_id,
                call_type="implementatieplanning_vraag",
            )
            raw = resp.content[0].text.strip()
        except Exception as e:
            logger.error(f"[ip-chat/vraag] Claude fout: {e}")
            raise HTTPException(500, f"AI-vraag mislukt: {str(e)}")

        start = raw.find("{")
        eind  = raw.rfind("}") + 1
        if start == -1 or eind == 0:
            raise HTTPException(500, "AI response bevat geen geldig JSON")
        try:
            data = json.loads(raw[start:eind])
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI gaf ongeldige JSON: {e}")

        return {
            "modus":   "vraag",
            "antwoord": data.get("antwoord", "Geen antwoord ontvangen."),
        }

    # ── Aanpas-modus: diff berekenen en toepassen ─────────────────────────────
    try:
        resp = await asyncio.to_thread(
            call_claude,
            messages=[{"role": "user", "content": user_message}],
            model=gekozen_model,
            max_tokens=4000,
            system=CHAT_SYSTEEM_PROMPT,
            log_usage=True,
            db=db,
            tender_id=tender_id,
            bureau_id=bureau_id,
            call_type="implementatieplanning_chat",
        )
        raw = resp.content[0].text.strip()
    except Exception as e:
        logger.error(f"[ip-chat] Claude fout: {e}")
        raise HTTPException(500, f"AI-chat mislukt: {str(e)}")

    start = raw.find("{")
    eind  = raw.rfind("}") + 1
    if start == -1 or eind == 0:
        raise HTTPException(500, "AI response bevat geen geldig JSON")
    try:
        diff = json.loads(raw[start:eind])
    except json.JSONDecodeError as e:
        logger.error(f"[ip-chat] JSON parse fout: {e}")
        raise HTTPException(500, f"AI gaf ongeldige JSON: {e}")

    toegepast = []
    for w in diff.get("wijzigingen", []):
        wtype = w.get("type")
        try:
            if wtype == "sectie_toevoegen":
                data  = dict(w["data"])
                taken = data.pop("taken", [])
                data.update({"tender_id": tender_id, "tenderbureau_id": bureau_id})
                res = db.table("implementatie_secties").insert(data).execute()
                nieuwe_id = res.data[0]["id"]
                for taak in taken:
                    taak.update({"sectie_id": nieuwe_id, "tender_id": tender_id,
                                 "tenderbureau_id": bureau_id})
                    taak["dagen"] = _bereken_dagen(taak.get("startdatum"), taak.get("einddatum")) \
                                    or taak.get("dagen")
                    db.table("implementatie_taken").insert(taak).execute()

            elif wtype == "taak_toevoegen":
                data = dict(w.get("data", {}))
                data.update({"sectie_id": w.get("sectie_id"), "tender_id": tender_id,
                              "tenderbureau_id": bureau_id})
                data["dagen"] = _bereken_dagen(data.get("startdatum"), data.get("einddatum")) \
                                 or data.get("dagen")
                db.table("implementatie_taken").insert(data).execute()

            elif wtype == "taak_wijzigen":
                data = dict(w.get("data", {}))
                if "startdatum" in data and "einddatum" in data:
                    data["dagen"] = _bereken_dagen(data["startdatum"], data["einddatum"])
                data["updated_at"] = _now()
                db.table("implementatie_taken").update(data).eq("id", w["taak_id"]).execute()

            elif wtype == "taak_verwijderen":
                db.table("implementatie_taken").delete().eq("id", w["taak_id"]).execute()

            elif wtype == "sectie_wijzigen":
                data = dict(w.get("data", {}))
                data["updated_at"] = _now()
                db.table("implementatie_secties").update(data).eq("id", w["sectie_id"]).execute()

            elif wtype == "sectie_verwijderen":
                db.table("implementatie_taken").delete().eq("sectie_id", w["sectie_id"]).execute()
                db.table("implementatie_secties").delete().eq("id", w["sectie_id"]).execute()

            else:
                logger.warning(f"[ip-chat] Onbekend wijzigingstype: {wtype}")
                continue

            toegepast.append(wtype)
        except Exception as e:
            logger.error(f"[ip-chat] Wijziging '{wtype}' mislukt: {e}")

    return {
        "modus":                 "aanpas",
        "samenvatting":          diff.get("samenvatting", "Planning bijgewerkt"),
        "wijzigingen_toegepast": len(toegepast),
        "types":                 toegepast,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/{tender_id}/export/excel")
async def export_excel(
    tender_id:    str,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl niet beschikbaar")

    planning  = _load_planning(db, tender_id)
    metadata  = planning["metadata"] or {}
    secties   = planning["secties"]
    projectnaam   = metadata.get("projectnaam", "Implementatieplanning")
    planstart_str = metadata.get("planstart")
    planeinde_str = metadata.get("planeinde")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Implementatieplanning"

    # Kolom breedte
    col_w = {"A": 6, "B": 42, "C": 22, "D": 15, "E": 12, "F": 12, "G": 7}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    # Weekkolommen berekenen
    week_labels = []
    if planstart_str and planeinde_str:
        try:
            ps = datetime.strptime(planstart_str, "%Y-%m-%d")
            pe = datetime.strptime(planeinde_str, "%Y-%m-%d")
            cur = ps
            while cur <= pe:
                week_labels.append(f"W{cur.strftime('%V')}\n{cur.strftime('%d/%m')}")
                cur += timedelta(weeks=1)
        except Exception:
            pass

    for i in range(len(week_labels)):
        ws.column_dimensions[get_column_letter(8 + i)].width = 4

    # Header rij
    DARK   = "1E293B"
    hfill  = PatternFill("solid", fgColor=DARK)
    hfont  = Font(bold=True, color="FFFFFF", size=9)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Nr", "Taak / Fase", "Verantwoordelijk", "Status", "Start", "Einde", "Dgn"] + week_labels
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill, cell.font, cell.alignment = hfill, hfont, halign
    ws.row_dimensions[1].height = 28

    STATUS_LABELS = {"open": "Open", "in_uitvoering": "In uitvoering", "afgerond": "Afgerond"}
    row = 2

    def _hex_fill(h):
        h = (h or "c7d2fe").lstrip("#")
        return PatternFill("solid", fgColor=h.upper())

    for sectie in secties:
        # Sectie header
        sfill  = _hex_fill(sectie.get("kleur"))
        sfont  = Font(bold=True, size=10)
        cell   = ws.cell(row=row, column=1, value=sectie["naam"])
        end_c  = get_column_letter(len(headers))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell.fill = sfill
        cell.font = sfont
        cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for taak in sectie.get("taken", []):
            ws.cell(row=row, column=1, value=taak.get("nummer",           ""))
            ws.cell(row=row, column=2, value=taak.get("naam",             ""))
            ws.cell(row=row, column=3, value=taak.get("verantwoordelijke", ""))
            ws.cell(row=row, column=4, value=STATUS_LABELS.get(taak.get("status", "open"), "Open"))
            ws.cell(row=row, column=5, value=taak.get("startdatum",       ""))
            ws.cell(row=row, column=6, value=taak.get("einddatum",        ""))
            ws.cell(row=row, column=7, value=taak.get("dagen") or "")

            # Gantt balk kolommen
            if week_labels and planstart_str and taak.get("startdatum") and taak.get("einddatum"):
                try:
                    ps = datetime.strptime(planstart_str, "%Y-%m-%d")
                    ts = datetime.strptime(taak["startdatum"], "%Y-%m-%d")
                    te = datetime.strptime(taak["einddatum"],  "%Y-%m-%d")
                    for wi in range(len(week_labels)):
                        w_start = ps + timedelta(weeks=wi)
                        w_end   = ps + timedelta(weeks=wi + 1)
                        if ts < w_end and te >= w_start:
                            ws.cell(row=row, column=8 + wi).fill = sfill
                except Exception:
                    pass

            ws.row_dimensions[row].height = 15
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode("utf-8")

    return {
        "bestandsnaam": f"Implementatieplanning_{projectnaam.replace(' ', '_')}.xlsx",
        "base64":       b64,
        "mimetype":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


# ── PDF export ────────────────────────────────────────────────────────────────

@router.post("/implementatieplanning/{tender_id}/export/pdf")
async def export_pdf(
    tender_id:    str,
    db:           Client = Depends(get_supabase_async),
    current_user: dict   = Depends(get_current_user),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as rl_canvas
    except ImportError:
        raise HTTPException(500, "reportlab niet beschikbaar — voer 'pip install reportlab' uit")

    planning      = _load_planning(db, tender_id)
    metadata      = planning["metadata"] or {}
    secties       = planning["secties"]
    projectnaam   = metadata.get("projectnaam",   "Implementatieplanning")
    opdrachtgever = metadata.get("opdrachtgever", "")
    planstart_str = metadata.get("planstart")
    planeinde_str = metadata.get("planeinde")

    buf   = io.BytesIO()
    pw, ph = landscape(A4)
    c      = rl_canvas.Canvas(buf, pagesize=(pw, ph))

    MG     = 15 * mm
    ROW_H  = 6.5 * mm

    # Bereken totaal projectdagen voor Gantt-fractie
    totaal_dagen = 1
    ps_dt = pe_dt = None
    if planstart_str and planeinde_str:
        try:
            ps_dt = datetime.strptime(planstart_str, "%Y-%m-%d")
            pe_dt = datetime.strptime(planeinde_str, "%Y-%m-%d")
            totaal_dagen = max(1, (pe_dt - ps_dt).days)
        except Exception:
            pass

    def _hex_rgb(h):
        h = (h or "c7d2fe").lstrip("#")
        if len(h) == 6:
            return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))
        return (0.78, 0.82, 0.996)

    # Kolombreedtes: [nr, naam, verantw, status, start, einde, dgn, gantt]
    tabel_w  = pw - 2 * MG
    gantt_w  = tabel_w * 0.42
    info_w   = tabel_w - gantt_w
    col_ws   = [
        info_w * 0.07,  # Nr
        info_w * 0.32,  # Naam
        info_w * 0.24,  # Verantw.
        info_w * 0.12,  # Status
        info_w * 0.11,  # Start
        info_w * 0.11,  # Einde
        info_w * 0.03,  # Dgn
        gantt_w,        # Gantt
    ]

    def _new_page():
        nonlocal y
        c.showPage()
        y = ph - MG
        # Herdruk kolomkoppen
        _draw_header_row()

    def _draw_cell(x, cy, text, w, bold=False, fill_rgb=None, text_color=(0, 0, 0)):
        if fill_rgb:
            c.setFillColorRGB(*fill_rgb)
            c.rect(x, cy - ROW_H, w, ROW_H, fill=1, stroke=0)
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.rect(x, cy - ROW_H, w, ROW_H, fill=0, stroke=1)
        c.setFillColorRGB(*text_color)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 7.5)
        c.drawString(x + 1.5 * mm, cy - ROW_H + 1.8 * mm, str(text or "")[:45])

    def _draw_header_row():
        nonlocal y
        lbls = ["Nr", "Taak / Fase", "Verantw.", "Status", "Start", "Einde", "Dgn", "Gantt"]
        x = MG
        for lbl, w in zip(lbls, col_ws):
            _draw_cell(x, y, lbl, w, bold=True,
                       fill_rgb=(0.12, 0.16, 0.24), text_color=(1, 1, 1))
            x += w
        y -= ROW_H

    # ── Paginakop ──────────────────────────────────────────────────────────
    y = ph - MG
    c.setFont("Helvetica-Bold", 13)
    c.setFillColorRGB(0.07, 0.09, 0.29)
    c.drawString(MG, y, projectnaam)

    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    if opdrachtgever:
        c.drawString(MG, y - 5 * mm, f"Opdrachtgever: {opdrachtgever}")
    if planstart_str and planeinde_str:
        c.drawRightString(pw - MG, y, f"Looptijd: {planstart_str}  –  {planeinde_str}")

    y -= 16 * mm
    _draw_header_row()

    STATUS_LABELS = {"open": "Open", "in_uitvoering": "Uitvoering", "afgerond": "Afgerond"}

    for sectie in secties:
        if y < MG + ROW_H * 3:
            _new_page()

        rgb = _hex_rgb(sectie.get("kleur"))
        x = MG
        for i, w in enumerate(col_ws):
            text = sectie["naam"] if i == 0 else ""
            # Sectieheader spans alle kolommen (visueel via kleur)
            _draw_cell(x, y, text if i == 0 else "", w,
                       bold=True, fill_rgb=rgb, text_color=(0.1, 0.1, 0.1))
            x += w
        # Tekst over breedte van eerste 7 kolommen
        c.setFillColorRGB(0.1, 0.1, 0.1)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(MG + 1.5 * mm, y - ROW_H + 1.8 * mm, sectie["naam"])
        y -= ROW_H

        for taak in sectie.get("taken", []):
            if y < MG + ROW_H:
                _new_page()

            cols_data = [
                taak.get("nummer",            ""),
                taak.get("naam",              ""),
                taak.get("verantwoordelijke", ""),
                STATUS_LABELS.get(taak.get("status", "open"), "Open"),
                taak.get("startdatum",        ""),
                taak.get("einddatum",         ""),
                str(taak.get("dagen") or ""),
                "",  # gantt placeholder
            ]
            x = MG
            for i, (text, w) in enumerate(zip(cols_data, col_ws)):
                if i < 7:
                    _draw_cell(x, y, text, w)
                else:
                    # Gantt cel achtergrond
                    c.setFillColorRGB(0.96, 0.97, 0.99)
                    c.rect(x, y - ROW_H, w, ROW_H, fill=1, stroke=0)
                    c.setStrokeColorRGB(0.85, 0.85, 0.85)
                    c.rect(x, y - ROW_H, w, ROW_H, fill=0, stroke=1)

                    # Gantt balk
                    if ps_dt and taak.get("startdatum") and taak.get("einddatum"):
                        try:
                            ts = datetime.strptime(taak["startdatum"], "%Y-%m-%d")
                            te = datetime.strptime(taak["einddatum"],  "%Y-%m-%d")
                            l_frac = max(0.0, (ts - ps_dt).days / totaal_dagen)
                            w_frac = max(0.005, (te - ts).days  / totaal_dagen)
                            bar_x  = x + l_frac * w
                            bar_w  = min(w_frac * w, w - l_frac * w)
                            pad    = 1.2 * mm
                            c.setFillColorRGB(*rgb)
                            c.rect(bar_x + pad, y - ROW_H + pad,
                                   max(1, bar_w - 2 * pad), ROW_H - 2 * pad,
                                   fill=1, stroke=0)
                        except Exception:
                            pass
                x += w

            y -= ROW_H

    c.save()
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode("utf-8")

    return {
        "bestandsnaam": f"Implementatieplanning_{projectnaam.replace(' ', '_')}.pdf",
        "base64":       b64,
        "mimetype":     "application/pdf",
    }
