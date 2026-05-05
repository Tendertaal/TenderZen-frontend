"""
Offerte Calculator API — TenderZen
Berekent uren en factuurbedragen voor het schrijven van een aanbesteding.
Gebaseerd op het bestaande Excel-model van Tendertaal.
"""
import asyncio
import base64
import io
import json
import logging
import re
import uuid as uuid_lib
from datetime import date, datetime
from typing import Optional, List

import fitz  # PyMuPDF

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from supabase import Client

from app.core.dependencies import get_current_user
from app.core.database import get_supabase_async
from app.services.anthropic_service import call_claude
from app.config import TOEGESTANE_MODELLEN, DEFAULT_AI_MODEL

OFFERTE_STORAGE_BUCKET = "ai-documents"
OFFERTE_STORAGE_PREFIX = "offerte-analyses"

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/offerte-calculator", tags=["Offerte Calculator"])


class AnalyseerRequest(BaseModel):
    model: Optional[str] = None

# ---------------------------------------------------------------------------
# Urenmodel — exact zoals Excel template Tendertaal
# ---------------------------------------------------------------------------

ACTIVITEITEN = [
    {"naam": "Verkoop gerelateerde activiteiten",      "uren": 2,    "comp": "percelen"},
    {"naam": "Voorbereiding: inlezen aanbestedingsstukken", "uren": 2, "comp": "percelen"},
    {"naam": "Strategische sessie/Kick-off",           "uren": 2,    "comp": "percelen"},
    {"naam": "Vragen NvI formuleren",                  "uren": 0.25, "comp": "vragen_nvi"},
    {"naam": "Interview(s) per gunningscriterium",     "uren": 1,    "comp": "sub_criteria"},
    {"naam": "Reviewsessies met klant (3 versies)",    "uren": 1,    "comp": "fixed_3"},
    {"naam": "Tussentijds overleg/communicatie",       "uren": 1,    "comp": "percelen"},
    {"naam": "Doornemen NvI stukken",                  "uren": 1,    "comp": "percelen"},
    {"naam": "Teksten versie 1 per pagina",            "uren": 2,    "comp": "paginas"},
    {"naam": "Teksten versie 2 per pagina",            "uren": 1,    "comp": "paginas"},
    {"naam": "Teksten versie 3 per pagina",            "uren": 1,    "comp": "paginas"},
    {"naam": "Bijlagen redigeren per pagina",          "uren": 1,    "comp": "bijlagen_redigeren"},
    {"naam": "Presentatie support",                    "uren": 2,    "comp": "presentatie"},
]

ANALYSE_PROMPT = """
Analyseer deze aanbesteding en extraheer de volgende variabelen voor een offerteberekening.

Retourneer UITSLUITEND dit JSON formaat, geen uitleg, geen markdown:
{
  "percelen": 1,
  "sub_criteria": 3,
  "paginas": 8,
  "bijlagen_redigeren": 1,
  "presentatie": false,
  "type_opdracht": "Diensten",
  "basisperiode": 2,
  "verlengopties": 2,
  "waarde_min": 400000,
  "waarde_max": 600000,
  "type_aanbesteding": "Meervoudig onderhands",
  "toelichting": "Samenvatting van de gemaakte keuzes",
  "toelichtingen": {
    "percelen": "Eén zin hoe dit geschat is",
    "sub_criteria": "Eén zin hoe dit geschat is",
    "paginas": "Eén zin hoe dit geschat is",
    "bijlagen_redigeren": "Eén zin hoe dit geschat is",
    "basisperiode": "Eén zin hoe dit geschat is",
    "verlengopties": "Eén zin hoe dit geschat is",
    "waarde_min": "Eén zin hoe dit geschat is",
    "waarde_max": "Eén zin hoe dit geschat is"
  }
}

Toelichting per veld:
- percelen: aantal percelen in de aanbesteding (default 1)
- sub_criteria: aantal gunningscriteria of sub-criteria
- paginas: geschat aantal A4 pagina's voor de inschrijving
- bijlagen_redigeren: aantal bijlagen die inhoudelijk uitgewerkt moeten worden
- presentatie: verwacht een presentatie als onderdeel van de procedure
- type_opdracht: Diensten / Leveringen / Werken
- basisperiode: basiscontractduur in jaren (exclusief verlengopties)
- verlengopties: optionele verlengperiode in jaren
- waarde_min: minimale geraamde opdrachtwaarde over de gehele looptijd in euros
- waarde_max: maximale geraamde opdrachtwaarde over de gehele looptijd in euros
- type_aanbesteding: Enkelvoudig onderhands / Meervoudig onderhands / Openbaar / Europees openbaar / Minicompetitie
- toelichtingen: per geschat veld een korte zin (max 2 zinnen) die uitlegt hoe de waarde is bepaald
"""


# ---------------------------------------------------------------------------
# Berekeningen
# ---------------------------------------------------------------------------

def bereken_uren(data: dict) -> dict:
    """Berekent uren op basis van variabelen — zelfde logica als Excel model.

    Gebruikt opgeslagen activiteiten_detail (met uren_indicatie en multiple_override)
    als die aanwezig zijn in data, anders de standaard ACTIVITEITEN lijst.
    """
    v = {
        "percelen":          data.get("percelen", 1) or 1,
        "sub_criteria":      data.get("sub_criteria", 0) or 0,
        "paginas":           data.get("paginas", 0) or 0,
        "bijlagen_redigeren":data.get("bijlagen_redigeren", 0) or 0,
        "vragen_nvi":        data.get("vragen_nvi", 0) or 0,
        "presentatie":       1 if data.get("presentatie") else 0,
        "fixed_3":           3,
        "fixed_1":           1,
    }

    # Gebruik opgeslagen activiteiten als die er zijn (met aangepaste uren en multipliers)
    opgeslagen = data.get("activiteiten_detail") or []
    if opgeslagen:
        opgeslagen_map = {a["naam"]: a for a in opgeslagen}
        bronlijst = []
        for a in ACTIVITEITEN:
            saved = opgeslagen_map.get(a["naam"])
            bronlijst.append({
                "naam":             a["naam"],
                "uren":             float(saved["uren_indicatie"]) if saved else a["uren"],
                "comp":             saved.get("component", a["comp"]) if saved else a["comp"],
                "multiple_override": saved.get("multiple_override") if saved else None,
            })
        # Voeg eventuele extra activiteiten toe die niet in de standaard lijst staan
        for saved in opgeslagen:
            if not any(a["naam"] == saved["naam"] for a in ACTIVITEITEN):
                bronlijst.append({
                    "naam":             saved["naam"],
                    "uren":             float(saved.get("uren_indicatie", 1)),
                    "comp":             saved.get("component", "fixed_1"),
                    "multiple_override": saved.get("multiple_override"),
                })
    else:
        bronlijst = [{"naam": a["naam"], "uren": a["uren"], "comp": a["comp"], "multiple_override": None}
                     for a in ACTIVITEITEN]

    totaal = 0.0
    activiteiten_detail = []
    for a in bronlijst:
        override = a.get("multiple_override")
        multiple = float(override) if override is not None else v.get(a["comp"], 0)
        uren = a["uren"] * multiple
        totaal += uren
        entry = {
            "naam":            a["naam"],
            "uren_indicatie":  a["uren"],
            "component":       a["comp"],
            "multiple":        multiple,
            "uren":            round(uren, 2),
        }
        if override is not None:
            entry["multiple_override"] = float(override)
        activiteiten_detail.append(entry)

    korting_pct = (
        (data.get("bekende_klant_pct", 0) or 0) +
        (data.get("zittende_partij_pct", 0) or 0)
    ) / 100
    uren_in_mindering = totaal * korting_pct
    uren_netto = totaal - uren_in_mindering
    uurtarief = data.get("uurtarief", 130) or 130
    bedrag_berekend = uren_netto * uurtarief

    return {
        "uren_berekend":        round(totaal, 2),
        "uren_in_mindering":    round(uren_in_mindering, 2),
        "uren_netto":           round(uren_netto, 2),
        "bedrag_berekend":      round(bedrag_berekend, 2),
        "activiteiten_detail":  activiteiten_detail,
    }


def bereken_factuur(data: dict) -> dict:
    """Berekent factuurbedragen en commissie — commissie alleen over tenderschrijven."""
    schrijven   = float(data.get("factuur_tenderschrijven", 0) or 0)
    management  = float(data.get("factuur_tendermanagement", 0) or 0)
    documenten  = float(data.get("factuur_tendercdocumenten", 0) or 0)
    grafisch    = float(data.get("factuur_grafisch_ontwerp", 0) or 0)
    totaal      = schrijven + management + documenten + grafisch

    commissie_pct = (data.get("commissie_pct", 10) or 10) / 100
    commissie = schrijven * commissie_pct  # alleen over tenderschrijven
    netto = totaal - commissie

    return {
        "factuur_totaal":   round(totaal, 2),
        "commissie_bedrag": round(commissie, 2),
        "netto_tendertaal": round(netto, 2),
    }


def extraheer_json(tekst: str) -> str:
    """Extraheert JSON uit Claude response, ook als die in markdown backticks zit."""
    if "```" in tekst:
        match = re.search(r'```(?:json)?\s*([\s\S]*?)```', tekst)
        if match:
            tekst = match.group(1).strip()
    start = tekst.find('{')
    end = tekst.rfind('}')
    if start != -1 and end != -1:
        return tekst[start:end + 1]
    return tekst


def genereer_excel(offerte: dict, uren_detail: list) -> bytes:
    """Genereert Excel bestand met dezelfde layout als het Tendertaal template."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    inschrijver = offerte.get("inschrijvende_partij") or "Offerte"
    ws.title = inschrijver[:31]  # Excel max 31 tekens voor sheetnaam

    # Stijlen
    header_font = Font(bold=True)
    paars_fill  = PatternFill("solid", fgColor="7C3AED")
    paars_font  = Font(color="FFFFFF", bold=True)
    groen_fill  = PatternFill("solid", fgColor="16A34A")
    groen_font  = Font(color="FFFFFF", bold=True)
    geel_fill   = PatternFill("solid", fgColor="FEF3C7")

    def bold(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = header_font

    # Header sectie
    ws['A1'] = 'Offerte Aangevraagd door:'
    ws['B1'] = offerte.get("aangemaakt_door_naam", "")
    ws['E1'] = 'Kwaliteit'
    ws['F1'] = (offerte.get("kwaliteit_weging") or 60) / 100

    ws['A2'] = 'Inschrijvende partij:'
    ws['B2'] = offerte.get("inschrijvende_partij", "")
    ws['E2'] = 'Prijs'
    ws['F2'] = (offerte.get("prijs_weging") or 40) / 100

    ws['A3'] = 'Aanbestedende dienst:'
    ws['B3'] = offerte.get("aanbestedende_dienst", "")

    ws['A4'] = 'Aanbesteding:'
    ws['B4'] = offerte.get("aanbesteding", "")

    ws['A5'] = 'Type aanbesteding:'
    ws['B5'] = offerte.get("type_aanbesteding", "")

    ws['A7'] = 'DEADLINE indienen'
    ws['A7'].font = header_font
    if offerte.get("deadline"):
        ws['B7'] = str(offerte["deadline"])

    # Variabelen sectie
    bold(9, 1, 'Variabelen:')
    bold(9, 2, 'Aantal:')
    bold(9, 5, 'Opdrachtwaarde:')

    ws['A10'] = 'Percelen';          ws['B10'] = offerte.get("percelen", 1)
    ws['E10'] = 'Type opdracht';     ws['F10'] = offerte.get("type_opdracht", "")
    ws['A12'] = 'Sub-criteria aantal'; ws['B12'] = offerte.get("sub_criteria", 0)
    ws['E12'] = 'Looptijd in j';     ws['F12'] = offerte.get("looptijd_jaar", 0)
    ws['A13'] = "Pagina's";          ws['B13'] = offerte.get("paginas", 0)
    ws['E13'] = 'Waarde';            ws['F13'] = offerte.get("waarde", 0)
    ws['A14'] = 'Bijlagen redigeren/uitwerken'; ws['B14'] = offerte.get("bijlagen_redigeren", 0)
    ws['A15'] = 'Zittende partij';   ws['B15'] = (offerte.get("zittende_partij_pct", 0) or 0) / 100
    ws['A16'] = 'Bekende klant';     ws['B16'] = (offerte.get("bekende_klant_pct", 0) or 0) / 100

    # Activiteiten tabel
    header_row = 18
    for i, h in enumerate(['Activiteit', '', '', '', '', '', '', 'Indicatie Uren', 'Component', 'Multiple', 'Uren']):
        c = ws.cell(row=header_row, column=i+1, value=h)
        c.font = header_font

    for i, act in enumerate(uren_detail):
        row = header_row + 1 + i
        ws.cell(row=row, column=1,  value=act['naam'])
        ws.cell(row=row, column=8,  value=act['uren_indicatie'])
        ws.cell(row=row, column=9,  value=act['component'])
        ws.cell(row=row, column=10, value=act['multiple'])
        ws.cell(row=row, column=11, value=act['uren'])

    # Resultaten sectie
    rc  = 10  # result column
    rs  = header_row + len(uren_detail) + 2

    bold(rs, rc, 'FACTUURBEDRAG KLANT')
    ws.cell(rs, rc).font = Font(bold=True, color="FFFFFF")
    ws.cell(rs, rc).fill = paars_fill

    rows_factuur = [
        ('Tenderschrijven',   'factuur_tenderschrijven'),
        ('Tendermanagement',  'factuur_tendermanagement'),
        ('Tendercdocumenten', 'factuur_tendercdocumenten'),
        ('Grafisch ontw.',    'factuur_grafisch_ontwerp'),
        ('Totaal',            'factuur_totaal'),
    ]
    for j, (label, key) in enumerate(rows_factuur):
        r = rs + 1 + j
        ws.cell(r, rc, label)
        ws.cell(r, rc + 1, offerte.get(key, 0) or 0)

    bold(rs + 8, rc, 'Tendertaal Offerte:')
    ws.cell(rs + 9,  rc, 'Uurtarief');               ws.cell(rs + 9,  rc + 1, offerte.get("uurtarief", 130))
    ws.cell(rs + 10, rc, 'Uren');                     ws.cell(rs + 10, rc + 1, offerte.get("uren_netto", 0))
    ws.cell(rs + 11, rc, 'Offerte bedrag Schrijven'); ws.cell(rs + 11, rc + 1, offerte.get("bedrag_berekend", 0))

    bold(rs + 13, rc, 'BEDRAG OPGENOMEN IN OFFERTE')
    ws.cell(rs + 13, rc + 1, offerte.get("factuur_tenderschrijven", 0) or 0)

    commissie_label = f"Verkoopcommissie {offerte.get('commissie_naam', 'Rick')}:"
    ws.cell(rs + 15, rc,     commissie_label)
    ws.cell(rs + 15, rc + 1, (offerte.get("commissie_pct", 10) or 10) / 100)
    ws.cell(rs + 15, rc + 2, offerte.get("commissie_bedrag", 0) or 0)

    c_netto = ws.cell(rs + 17, rc, 'NETTO BLIJFT DIT BEDRAG OVER VOOR TENDERTAAL:')
    c_netto.font = groen_font
    c_netto.fill = groen_fill
    c_netto2 = ws.cell(rs + 17, rc + 1, offerte.get("netto_tendertaal", 0) or 0)
    c_netto2.font = groen_font
    c_netto2.fill = groen_fill

    # Kolombreedtes
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['J'].width = 28
    ws.column_dimensions['K'].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Pydantic modellen
# ---------------------------------------------------------------------------

class OfferteAanmaken(BaseModel):
    tender_id:            Optional[str] = None
    inschrijvende_partij: Optional[str] = None
    aanbestedende_dienst: Optional[str] = None
    aanbesteding:         Optional[str] = None
    type_aanbesteding:    Optional[str] = None
    deadline:             Optional[str] = None
    kwaliteit_weging:     Optional[int] = 60
    prijs_weging:         Optional[int] = 40


class OfferteUpdate(BaseModel):
    inschrijvende_partij:    Optional[str] = None
    aanbestedende_dienst:    Optional[str] = None
    aanbesteding:            Optional[str] = None
    type_aanbesteding:       Optional[str] = None
    deadline:                Optional[str] = None
    kwaliteit_weging:        Optional[int] = None
    prijs_weging:            Optional[int] = None
    type_opdracht:           Optional[str] = None
    looptijd_jaar:           Optional[float] = None
    basisperiode:            Optional[float] = None
    verlengopties:           Optional[float] = None
    waarde:                  Optional[float] = None
    waarde_min:              Optional[float] = None
    waarde_max:              Optional[float] = None
    ai_toelichtingen:        Optional[dict] = None
    percelen:                Optional[int] = None
    sub_criteria:            Optional[int] = None
    paginas:                 Optional[int] = None
    bijlagen:                Optional[int] = None
    vragen_nvi:              Optional[int] = None
    bijlagen_redigeren:      Optional[int] = None
    presentatie:             Optional[bool] = None
    bekende_klant_pct:       Optional[int] = None
    zittende_partij_pct:     Optional[int] = None
    uurtarief:               Optional[int] = None
    factuur_tenderschrijven: Optional[float] = None
    factuur_tendermanagement:Optional[float] = None
    factuur_tendercdocumenten:Optional[float] = None
    factuur_grafisch_ontwerp:Optional[float] = None
    commissie_naam:             Optional[str] = None
    commissie_pct:              Optional[int] = None
    tarief_tenderschrijven:     Optional[float] = None
    tarief_tendermanagement:    Optional[float] = None
    tarief_grafisch_per_pagina: Optional[float] = None
    korting_tenderschrijven:    Optional[int] = None
    korting_tendermanagement:   Optional[int] = None
    korting_grafisch:           Optional[int] = None
    commissie_basis:            Optional[str] = None
    status:                     Optional[str] = None
    notities:                   Optional[str] = None
    schrijver_type:             Optional[str] = None
    schrijver_user_id:          Optional[str] = None
    schrijver_naam:             Optional[str] = None
    manager_user_id:            Optional[str] = None
    manager_naam:               Optional[str] = None
    grafisch_user_id:           Optional[str] = None
    grafisch_naam:              Optional[str] = None
    inhuur_tarief_schrijven:    Optional[float] = None
    netto_include_schrijven:    Optional[bool] = None
    netto_include_management:   Optional[bool] = None
    netto_include_documenten:   Optional[bool] = None
    netto_include_grafisch:     Optional[bool] = None
    netto_include_inhuur:       Optional[bool] = None
    netto_include_commissie:    Optional[bool] = None
    activiteiten_detail:        Optional[list] = None


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("")
async def lijst_offerte_calculaties(
    bureau_id: Optional[str] = Query(None),
    tender_id: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Haal alle offerte-calculaties op. Super-admin zonder bureau_id ziet alles."""
    is_super_admin   = current_user.get("is_super_admin", False)
    tenderbureau_id  = bureau_id or current_user.get("tenderbureau_id")

    q = db.table("offerte_calculaties").select("*").order("created_at", desc=True)

    if is_super_admin and not bureau_id:
        pass  # super-admin zonder filter → alle offertes van alle bureaus
    elif tenderbureau_id:
        q = q.eq("tenderbureau_id", tenderbureau_id)
    else:
        raise HTTPException(status_code=400, detail="Geen bureau_id beschikbaar")

    if tender_id:
        q = q.eq("tender_id", tender_id)

    res = q.execute()
    return {"calculaties": res.data or []}


@router.get("/{offerte_id}")
async def haal_offerte_op(
    offerte_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    def _query(client):
        res = client.table("offerte_calculaties").select("*").eq("id", offerte_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Offerte niet gevonden")
        return res.data[0]

    try:
        offerte = _query(db)
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"GET offerte timeout/fout ({e}), retry met verse client")
        try:
            from app.core.database import get_supabase_admin
            offerte = _query(get_supabase_admin())
        except HTTPException:
            raise
        except Exception as e2:
            raise HTTPException(status_code=500, detail=f"Database timeout: {str(e2)}")

    uren = bereken_uren(offerte)
    return {"offerte": offerte, "activiteiten_detail": uren["activiteiten_detail"]}


@router.post("")
async def maak_offerte_aan(
    body: OfferteAanmaken,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    bureau_id = current_user.get("tenderbureau_id")
    if not bureau_id:
        raise HTTPException(status_code=400, detail="Geen bureau_id gevonden voor gebruiker")

    data = body.dict(exclude_none=True)
    data["tenderbureau_id"] = bureau_id
    data["aangemaakt_door"]  = current_user.get("id")

    # Server-side uren berekenen op basis van defaults
    uren_result = bereken_uren(data)
    data.update({
        "uren_berekend":     uren_result["uren_berekend"],
        "uren_in_mindering": uren_result["uren_in_mindering"],
        "uren_netto":        uren_result["uren_netto"],
        "bedrag_berekend":   uren_result["bedrag_berekend"],
    })

    # Factuur berekenen (defaults = 0)
    factuur_result = bereken_factuur(data)
    data.update(factuur_result)

    res = db.table("offerte_calculaties").insert(data).execute()
    if not res.data:
        raise HTTPException(status_code=500, detail="Aanmaken mislukt")
    return {"offerte": res.data[0], "activiteiten_detail": uren_result["activiteiten_detail"]}


@router.put("/{offerte_id}")
async def update_offerte(
    offerte_id: str,
    body: OfferteUpdate,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    try:
        bestaand_res = db.table("offerte_calculaties").select("*").eq("id", offerte_id).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database fout bij ophalen: {str(e)}")

    if not bestaand_res.data:
        raise HTTPException(status_code=404, detail="Offerte niet gevonden")

    bestaand = bestaand_res.data[0]
    update_data = {k: v for k, v in body.dict(exclude_none=True).items()}

    # Zorg dat INTEGER-kolommen echte integers zijn (geen floats zoals 1.0)
    _int_kolommen = {
        'looptijd_jaar', 'kwaliteit_weging', 'prijs_weging', 'percelen',
        'sub_criteria', 'paginas', 'bijlagen', 'vragen_nvi', 'bijlagen_redigeren',
        'bekende_klant_pct', 'zittende_partij_pct', 'uurtarief', 'commissie_pct',
        'korting_tenderschrijven', 'korting_tendermanagement', 'korting_grafisch',
    }
    for veld in _int_kolommen:
        if veld in update_data and update_data[veld] is not None:
            try:
                update_data[veld] = int(float(update_data[veld]))
            except (ValueError, TypeError):
                pass

    # Merge met bestaande data voor berekeningen
    merged = {**bestaand, **update_data}

    # Herbereken uren en factuur server-side
    uren_result    = bereken_uren(merged)
    factuur_result = bereken_factuur(merged)
    update_data.update({
        "uren_berekend":     uren_result["uren_berekend"],
        "uren_in_mindering": uren_result["uren_in_mindering"],
        "uren_netto":        uren_result["uren_netto"],
        "bedrag_berekend":   uren_result["bedrag_berekend"],
        **factuur_result,
    })

    try:
        res = db.table("offerte_calculaties").update(update_data).eq("id", offerte_id).execute()
    except Exception as e:
        logger.error(f"Opslaan offerte {offerte_id} mislukt: {e}")
        raise HTTPException(status_code=500, detail=f"Opslaan mislukt: {str(e)}")

    if not res.data:
        raise HTTPException(status_code=500, detail="Opslaan mislukt: geen data terug van database")
    return {"offerte": res.data[0], "activiteiten_detail": uren_result["activiteiten_detail"]}


@router.delete("/{offerte_id}")
async def verwijder_offerte(
    offerte_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    db.table("offerte_calculaties").delete().eq("id", offerte_id).execute()
    return {"ok": True}


@router.post("/tender/{tender_id}/nieuw")
async def nieuw_vanuit_tender(
    tender_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Maak een nieuwe offerte-calculatie aan op basis van tenderdata."""
    # Haal tenderdata op vóór de bureau_id check — super-admin heeft geen tenderbureau_id
    # op current_user, maar de tender zelf heeft er altijd één.
    tender_res = db.table("tenders").select(
        "id,naam,opdrachtgever,aanbestedende_dienst,deadline_indiening,"
        "aanbestedingsprocedure,minimum_bedrag,maximum_bedrag,tenderbureau_id"
    ).eq("id", tender_id).execute()

    if not tender_res.data:
        raise HTTPException(status_code=404, detail="Tender niet gevonden")

    tender = tender_res.data[0]
    bureau_id = tender.get("tenderbureau_id") or current_user.get("tenderbureau_id")
    if not bureau_id:
        raise HTTPException(status_code=400, detail="Geen tenderbureau_id beschikbaar")

    deadline_str = None
    if tender.get("deadline_indiening"):
        try:
            deadline_str = str(tender["deadline_indiening"])[:10]
        except Exception:
            pass

    data = {
        "tenderbureau_id":    bureau_id,
        "aangemaakt_door":    current_user.get("id"),
        "tender_id":          tender_id,
        "aanbesteding":       tender.get("naam", ""),
        "aanbestedende_dienst": tender.get("opdrachtgever") or tender.get("aanbestedende_dienst", ""),
        "deadline":           deadline_str,
        "type_aanbesteding":  tender.get("aanbestedingsprocedure", ""),
        "waarde_min":         float(tender.get("minimum_bedrag") or 0) or None,
        "waarde_max":         float(tender.get("maximum_bedrag") or 0) or None,
        "waarde":             float(tender.get("maximum_bedrag") or tender.get("minimum_bedrag") or 0) or None,
    }
    uren_result    = bereken_uren(data)
    factuur_result = bereken_factuur(data)
    data.update({
        "uren_berekend":     uren_result["uren_berekend"],
        "uren_in_mindering": uren_result["uren_in_mindering"],
        "uren_netto":        uren_result["uren_netto"],
        "bedrag_berekend":   uren_result["bedrag_berekend"],
        **factuur_result,
    })

    res = db.table("offerte_calculaties").insert(data).execute()
    if not res.data:
        raise HTTPException(status_code=500, detail="Aanmaken mislukt")
    return {"offerte": res.data[0], "activiteiten_detail": uren_result["activiteiten_detail"]}


@router.post("/{offerte_id}/analyseer")
async def analyseer_met_ai(
    offerte_id: str,
    request: Optional[AnalyseerRequest] = None,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Analyseer de gekoppelde aanbestedingsdocumenten met Claude en extraheer variabelen."""
    req = request or AnalyseerRequest()
    gekozen_model = req.model or DEFAULT_AI_MODEL
    if gekozen_model not in TOEGESTANE_MODELLEN:
        raise HTTPException(status_code=400, detail=f"Ongeldig model: '{gekozen_model}'. Toegestane waarden: {sorted(TOEGESTANE_MODELLEN)}")

    offerte_res = db.table("offerte_calculaties").select("*").eq("id", offerte_id).execute()
    if not offerte_res.data:
        raise HTTPException(status_code=404, detail="Offerte niet gevonden")
    offerte = offerte_res.data[0]

    # Bouw aanbestedingstekst op
    aanbestedingstekst = ""

    # 1. Probeer geüploade offerte-documenten te lezen (PDF-tekst extraheren)
    offerte_docs_res = db.table("offerte_documenten") \
        .select("storage_path, original_file_name") \
        .eq("offerte_id", offerte_id) \
        .eq("gebruikt_voor_analyse", True) \
        .limit(5).execute()

    for doc in (offerte_docs_res.data or []):
        storage_path = doc.get("storage_path", "")
        try:
            pad = storage_path[len(f"{OFFERTE_STORAGE_PREFIX}/"):] if storage_path.startswith(f"{OFFERTE_STORAGE_PREFIX}/") else storage_path
            bestand_bytes = db.storage.from_(OFFERTE_STORAGE_BUCKET).download(pad)
            pdf_doc = fitz.open(stream=bestand_bytes, filetype="pdf")
            tekst = "\n".join(pagina.get_text() for pagina in pdf_doc)
            if tekst.strip():
                aanbestedingstekst += f"\n\n--- {doc['original_file_name']} ---\n{tekst[:6000]}"
        except Exception as e:
            logger.warning(f"PDF extractie mislukt voor {storage_path}: {e}")
            aanbestedingstekst += f"\nDocument: {doc.get('original_file_name', '')}"

    # 2. Tender-documenten als aanvullende context
    if offerte.get("tender_id"):
        docs_res = db.table("tender_documents") \
            .select("id,original_file_name,file_name,storage_path,document_type") \
            .eq("tender_id", offerte["tender_id"]) \
            .eq("is_deleted", False).limit(5).execute()
        for doc in (docs_res.data or []):
            storage_path = doc.get("storage_path", "")
            naam = doc.get("original_file_name") or doc.get("file_name", "")
            if storage_path:
                try:
                    bucket = "smart-imports" if storage_path.startswith("smart-imports/") else OFFERTE_STORAGE_BUCKET
                    pad = storage_path[len("smart-imports/"):] if storage_path.startswith("smart-imports/") else storage_path
                    bestand_bytes = db.storage.from_(bucket).download(pad)
                    pdf_doc = fitz.open(stream=bestand_bytes, filetype="pdf")
                    tekst = "\n".join(pagina.get_text() for pagina in pdf_doc)
                    if tekst.strip():
                        aanbestedingstekst += f"\n\n--- {naam} ---\n{tekst[:4000]}"
                        continue
                except Exception:
                    pass
            if naam:
                aanbestedingstekst += f"\nDocument: {naam}"

    # 3. Fallback: gebruik velden van de offerte zelf
    if not aanbestedingstekst.strip():
        aanbestedingstekst = "\n".join(filter(None, [
            offerte.get("aanbesteding", ""),
            offerte.get("aanbestedende_dienst", ""),
            offerte.get("type_aanbesteding", ""),
            offerte.get("omschrijving", ""),
        ]))

    if not aanbestedingstekst.strip():
        raise HTTPException(
            status_code=422,
            detail="Geen aanbestedingsdocumenten gevonden. Upload eerst documenten of vul de aanbestedingsvelden handmatig in."
        )

    gebruiker_bericht = f"Aanbesteding context:\n{aanbestedingstekst[:8000]}"
    messages = [{"role": "user", "content": gebruiker_bericht}]

    try:
        resp = await asyncio.to_thread(
            call_claude,
            messages=messages,
            system=ANALYSE_PROMPT,
            model=gekozen_model,
            max_tokens=1024,
            log_usage=False,
        )
        data = json.loads(extraheer_json(resp.content[0].text))
    except Exception as e:
        logger.error(f"AI-analyse mislukt: {e}")
        raise HTTPException(status_code=500, detail=f"AI-analyse mislukt: {str(e)}")

    # Uren herberekenen met AI-variabelen
    merged = {**offerte, **data}
    uren_result    = bereken_uren(merged)
    factuur_result = bereken_factuur(merged)

    update = {
        "percelen":           int(data.get("percelen", 1)),
        "sub_criteria":       int(data.get("sub_criteria", 0)),
        "paginas":            int(data.get("paginas", 0)),
        "bijlagen":           int(data.get("bijlagen", 0)),
        "vragen_nvi":         int(data.get("vragen_nvi", 0)),
        "bijlagen_redigeren": int(data.get("bijlagen_redigeren", 0)),
        "presentatie":        bool(data.get("presentatie", False)),
        "type_opdracht":      data.get("type_opdracht", ""),
        "basisperiode":       float(data.get("basisperiode", 0)) if data.get("basisperiode") else None,
        "verlengopties":      float(data.get("verlengopties", 0)) if data.get("verlengopties") else 0.0,
        "looptijd_jaar":      (float(data.get("basisperiode", 0)) or 0) + (float(data.get("verlengopties", 0)) or 0) or (float(data.get("looptijd_jaar", 0)) if data.get("looptijd_jaar") else None),
        "waarde":             float(data.get("waarde", 0)) if data.get("waarde") else None,
        "waarde_min":         float(data.get("waarde_min", 0)) if data.get("waarde_min") else None,
        "waarde_max":         float(data.get("waarde_max", 0)) if data.get("waarde_max") else None,
        "type_aanbesteding":  data.get("type_aanbesteding", offerte.get("type_aanbesteding", "")),
        "ai_geanalyseerd":    True,
        "ai_analyse_json":    data,
        "ai_toelichtingen":   data.get("toelichtingen") or {},
        **uren_result,
        **factuur_result,
    }
    update.pop("activiteiten_detail", None)

    db.table("offerte_calculaties").update(update).eq("id", offerte_id).execute()

    # Haal bijgewerkte offerte op — zodat frontend geen extra GET hoeft te doen
    bijgewerkt_res = db.table("offerte_calculaties").select("*").eq("id", offerte_id).execute()
    bijgewerkte_offerte = bijgewerkt_res.data[0] if bijgewerkt_res.data else None

    return {
        "offerte":             bijgewerkte_offerte,
        "variabelen":          data,
        "activiteiten_detail": uren_result["activiteiten_detail"],
        "toelichting":         data.get("toelichting", ""),
        "toelichtingen":       data.get("toelichtingen") or {},
    }


# ---------------------------------------------------------------------------
# Offerte documenten — upload, lijst, preview-url
# ---------------------------------------------------------------------------

@router.post("/{offerte_id}/documenten")
async def upload_offerte_document(
    offerte_id: str,
    bestanden: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Upload één of meer documenten bij een offerte (worden gebruikt bij AI-analyse)."""
    offerte_res = db.table("offerte_calculaties").select("id, tenderbureau_id").eq("id", offerte_id).execute()
    if not offerte_res.data:
        raise HTTPException(status_code=404, detail="Offerte niet gevonden")
    bureau_id = offerte_res.data[0].get("tenderbureau_id", "")

    opgeslagen = []
    for bestand in bestanden:
        inhoud = await bestand.read()
        bestand_id = str(uuid_lib.uuid4())
        storage_pad = f"{OFFERTE_STORAGE_PREFIX}/{bureau_id}/{offerte_id}/{bestand_id}_{bestand.filename}"

        try:
            db.storage.from_(OFFERTE_STORAGE_BUCKET).upload(
                path=storage_pad,
                file=inhoud,
                file_options={"content-type": bestand.content_type or "application/pdf"},
            )
        except Exception as e:
            logger.error(f"Storage upload mislukt: {e}")
            raise HTTPException(status_code=500, detail=f"Upload mislukt: {bestand.filename}")

        doc_res = db.table("offerte_documenten").insert({
            "offerte_id":          offerte_id,
            "tenderbureau_id":     bureau_id,
            "original_file_name":  bestand.filename,
            "storage_path":        storage_pad,
            "file_type":           bestand.content_type,
            "file_size":           len(inhoud),
            "gebruikt_voor_analyse": True,
        }).execute()

        opgeslagen.append(doc_res.data[0] if doc_res.data else {"original_file_name": bestand.filename})

    return {"documenten": opgeslagen}


@router.get("/{offerte_id}/documenten")
async def get_offerte_documenten(
    offerte_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Gecombineerde lijst: tender_documents van de gekoppelde tender + eigen offerte_documenten."""
    # Eigen geüploade offerte-documenten
    eigen_res = db.table("offerte_documenten") \
        .select("*") \
        .eq("offerte_id", offerte_id) \
        .order("created_at") \
        .execute()
    eigen = [{"source": "offerte", **d} for d in (eigen_res.data or [])]

    # Tender-documenten via de gekoppelde tender (geen dubbele opslag)
    offerte_res = db.table("offerte_calculaties").select("tender_id").eq("id", offerte_id).execute()
    tender_id = (offerte_res.data or [{}])[0].get("tender_id")

    tender = []
    if tender_id:
        tender_res = db.table("tender_documents") \
            .select("id, original_file_name, file_name, file_type, file_size, uploaded_at, storage_path") \
            .eq("tender_id", tender_id) \
            .eq("is_deleted", False) \
            .order("uploaded_at") \
            .execute()
        tender = [{
            "source":             "tender",
            "created_at":         d.get("uploaded_at"),
            "original_file_name": d.get("original_file_name") or d.get("file_name", ""),
            **d,
        } for d in (tender_res.data or [])]

    return tender + eigen


@router.get("/{offerte_id}/documenten/{document_id}/preview-url")
async def get_offerte_document_preview_url(
    offerte_id: str,
    document_id: str,
    source: str = Query(default="offerte"),
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Signed URL voor een document — werkt voor zowel offerte_documenten als tender_documents."""
    if source == "tender":
        doc_res = db.table("tender_documents") \
            .select("storage_path, original_file_name, file_name, file_type") \
            .eq("id", document_id) \
            .execute()
        if not doc_res.data:
            raise HTTPException(status_code=404, detail="Tender-document niet gevonden")
        doc = doc_res.data[0]
        storage_path = doc.get("storage_path", "")
        naam = doc.get("original_file_name") or doc.get("file_name", "")

        if storage_path.startswith("smart-imports/"):
            bucket = "smart-imports"
            pad    = storage_path[len("smart-imports/"):]
        else:
            bucket = OFFERTE_STORAGE_BUCKET
            pad    = storage_path
    else:
        doc_res = db.table("offerte_documenten") \
            .select("storage_path, original_file_name, file_type") \
            .eq("id", document_id) \
            .eq("offerte_id", offerte_id) \
            .execute()
        if not doc_res.data:
            raise HTTPException(status_code=404, detail="Document niet gevonden")
        doc = doc_res.data[0]
        naam = doc.get("original_file_name", "")
        storage_path = doc.get("storage_path", "")
        bucket = OFFERTE_STORAGE_BUCKET
        pad    = storage_path[len(f"{OFFERTE_STORAGE_PREFIX}/"):] if storage_path.startswith(f"{OFFERTE_STORAGE_PREFIX}/") else storage_path

    try:
        signed = db.storage.from_(bucket).create_signed_url(pad, 3600)
        signed_url = signed.get("signedURL") or signed.get("signedUrl") or ""
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Signed URL genereren mislukt: {e}")

    if not signed_url:
        raise HTTPException(status_code=500, detail="Kon geen toegangs-URL genereren")

    return {
        "url":       signed_url,
        "file_name": naam,
        "file_type": doc.get("file_type"),
    }


@router.delete("/{offerte_id}/documenten/{document_id}")
async def verwijder_offerte_document(
    offerte_id: str,
    document_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Verwijder een document bij een offerte (storage + database)."""
    doc_res = db.table("offerte_documenten") \
        .select("storage_path") \
        .eq("id", document_id) \
        .eq("offerte_id", offerte_id) \
        .execute()

    if not doc_res.data:
        raise HTTPException(status_code=404, detail="Document niet gevonden")

    storage_path = doc_res.data[0].get("storage_path", "")
    pad = storage_path[len(f"{OFFERTE_STORAGE_PREFIX}/"):] if storage_path.startswith(f"{OFFERTE_STORAGE_PREFIX}/") else storage_path

    try:
        db.storage.from_(OFFERTE_STORAGE_BUCKET).remove([pad])
    except Exception as e:
        logger.warning(f"Storage verwijdering mislukt: {e}")

    db.table("offerte_documenten").delete().eq("id", document_id).execute()
    return {"ok": True}


@router.post("/{offerte_id}/export-excel")
async def export_excel(
    offerte_id: str,
    current_user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_async),
):
    """Genereer Excel export van de offerte-calculatie."""
    offerte_res = db.table("offerte_calculaties").select("*").eq("id", offerte_id).execute()
    if not offerte_res.data:
        raise HTTPException(status_code=404, detail="Offerte niet gevonden")
    offerte = offerte_res.data[0]

    # Naam van aanmaker ophalen
    user_res = db.table("users").select("naam").eq("id", current_user.get("id")).execute()
    aangemaakt_door_naam = (user_res.data[0]["naam"] if user_res.data else "") or ""
    offerte["aangemaakt_door_naam"] = aangemaakt_door_naam

    uren = bereken_uren(offerte)
    excel_bytes = genereer_excel(offerte, uren["activiteiten_detail"])

    partij = (offerte.get("inschrijvende_partij") or "Offerte").replace(" ", "_")[:20]
    datum  = datetime.now().strftime("%Y%m%d")
    bestandsnaam = f"Offerte_{partij}_{datum}.xlsx"

    return {
        "bestandsnaam": bestandsnaam,
        "base64":       base64.b64encode(excel_bytes).decode("utf-8"),
    }
